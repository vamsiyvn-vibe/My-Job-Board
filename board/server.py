import socket
# Force IPv4 only to prevent connection hangs on misconfigured IPv6 environments
old_getaddrinfo = socket.getaddrinfo
def new_getaddrinfo(*args, **kwargs):
    responses = old_getaddrinfo(*args, **kwargs)
    return [r for r in responses if r[0] == socket.AF_INET]
socket.getaddrinfo = new_getaddrinfo

import os
import sys
import requests
import datetime
import traceback
import subprocess
import json
import threading
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPTS_DIR = r"G:\My Drive\.Agents\1_Projects\2026 Job Search\scripts"
if SCRIPTS_DIR not in sys.path:
    sys.path.append(SCRIPTS_DIR)

TRACKER_PATH = r"G:\My Drive\.Agents\1_Projects\2026 Job Search\Job_Leads_Tracker.xlsx"
CONFIG_PATH = r"G:\My Drive\.Agents\1_Projects\2026 Job Search\board\config.json"

app = Flask(__name__, static_url_path='', static_folder='.')
CORS(app)

# Global variables for async background scanning
is_scanning = False
auto_scan_triggered = False
scan_lock = threading.Lock()

# In-memory cache for jobs to avoid slow Excel loading on every request
JOBS_CACHE = {
    "jobs": [],
    "last_scanned": "",
    "last_loaded_mtime": 0
}

def load_jobs_from_excel():
    global JOBS_CACHE
    try:
        if not os.path.exists(TRACKER_PATH):
            return [], "Never"
            
        mtime = os.path.getmtime(TRACKER_PATH)
        # Reload cache if file is newer on disk or if cache is empty
        if mtime > JOBS_CACHE["last_loaded_mtime"] or not JOBS_CACHE["jobs"]:
            print(f"Loading jobs from Excel (mtime: {mtime})...")
            wb = openpyxl.load_workbook(TRACKER_PATH)
            if "Job Leads" in wb.sheetnames:
                ws = wb["Job Leads"]
                jobs = []
                # Read from row 2
                for r_idx in range(2, ws.max_row + 1):
                    company = ws.cell(row=r_idx, column=3).value
                    role = ws.cell(row=r_idx, column=4).value
                    
                    if not company or not role:
                        continue
                        
                    status = ws.cell(row=r_idx, column=7).value
                    if not status:
                        status = "Lead"
                        
                    url_cell = ws.cell(row=r_idx, column=8)
                    url = url_cell.hyperlink.target if url_cell.hyperlink else url_cell.value
                    
                    salary = ws.cell(row=r_idx, column=11).value or "N/A"
                    app_status = ws.cell(row=r_idx, column=12).value or "Not Applied"
                    app_outcome = ws.cell(row=r_idx, column=13).value or "Active / Pending"
                    resume_url = ws.cell(row=r_idx, column=14).value or ""
                    archive_reason = ws.cell(row=r_idx, column=15).value or ""
                    
                    jobs.append({
                        "id": r_idx,
                        "select": ws.cell(row=r_idx, column=1).value or "[ ]",
                        "cohort": ws.cell(row=r_idx, column=2).value or "Other",
                        "company": company,
                        "role": role,
                        "location": ws.cell(row=r_idx, column=5).value or "N/A",
                        "key_focus": ws.cell(row=r_idx, column=6).value or "",
                        "status": status,
                        "url": url,
                        "date_added": ws.cell(row=r_idx, column=9).value or "",
                        "notes": ws.cell(row=r_idx, column=10).value or "",
                        "salary": salary,
                        "app_status": app_status,
                        "app_outcome": app_outcome,
                        "resume_url": resume_url,
                        "archive_reason": archive_reason
                    })
                JOBS_CACHE["jobs"] = jobs
                JOBS_CACHE["last_scanned"] = datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %I:%M %p')
                JOBS_CACHE["last_loaded_mtime"] = mtime
                print(f"Loaded {len(jobs)} jobs from Excel into cache.")
            else:
                print("Sheet 'Job Leads' not found in spreadsheet.")
    except Exception as e:
        print("Error loading jobs from Excel file:", e)
        traceback.print_exc()
        # On error (e.g. file sharing lock), we retain the existing cache
    return JOBS_CACHE["jobs"], JOBS_CACHE["last_scanned"]

INTERVIEW_INSIGHTS = {
    "google": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Logistical check, resume walkthrough, and introductory behavioral screening questions."},
            {"title": "2. Product Sense Screen (45m)", "desc": "Product case interview led by a senior PM. Focuses on user-centric design, creative prioritization, and structured problem-solving."},
            {"title": "3. Final Onsite Loop (4-5 rounds)", "desc": "Rounds covering Product Design (user-first case), Analytical Thinking (metric definition/diagnostic), Product Strategy (long-term vision/monetization), and Leadership/Googleyness."}
        ],
        "tips": [
            "Use design frameworks like CIRCLES but tailor them dynamically to avoid sounding robotic or overly template-driven.",
            "Align all product recommendations with Google's core mission to make information universally accessible and useful.",
            "Think about extreme scale. Be ready to explain how solutions adapt to support billions of diverse daily users."
        ]
    },
    "microsoft": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Review of background, general motivation, and basic behavioral alignment check."},
            {"title": "2. Hiring Manager Screen (45m)", "desc": "Conversational screen evaluating product sense, role-specific strategy, and team cultural fit."},
            {"title": "3. Onsite Loop (3-6 rounds)", "desc": "Deep dive interviews covering Product Design & Strategy, Execution & Prioritization (data-driven decisions), and Behavioral Scenarios (collaboration/growth mindset)."}
        ],
        "tips": [
            "Demonstrate a growth mindset (a core Microsoft value) by discussing how you learn from failures and feedback.",
            "Focus on enterprise customer empathy. Understand multi-stakeholder needs, security compliance, and migration paths.",
            "Be highly structured: break complex, ambiguous enterprise/platform problems into logical phases."
        ]
    },
    "apple": {
        "stages": [
            {"title": "1. Recruiter Screen", "desc": "Initial logistics check, review of background, and compensation expectations."},
            {"title": "2. Hiring Manager Round", "desc": "Introductory conversation focusing on experience, team fit, and the crucial 'Why Apple?' assessment."},
            {"title": "3. Screening Rounds", "desc": "1-2 rounds focusing on product sense (situational design) and technical/domain depth depending on the team's area."},
            {"title": "4. Onsite Loop (4-5 rounds)", "desc": "Back-to-back rounds covering Apple-specific Product Sense & Strategy, Technical/Architectural Depth (trade-offs), and Culture/Behavioral alignment."}
        ],
        "tips": [
            "Prepare a highly authentic answer for 'Why Apple?' that aligns with their design philosophy, simplicity, and privacy values.",
            "Apple's PM loop is highly decentralized. Research your specific team's product lines and domain challenges deeply.",
            "Incorporate Apple's principles into your cases: prioritize user simplicity, privacy protection, and ecosystem continuity."
        ]
    },
    "amazon": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Basic alignment on background, logistics, and resume verification."},
            {"title": "2. Written Exercise (1-2 pages)", "desc": "Take-home written response to a behavioral prompt, structured in narrative STAR format, submitted within 48 hours."},
            {"title": "3. Phone Screen (60m)", "desc": "1-2 video rounds with PMs focusing on behavioral scenarios (Leadership Principles) and product design case questions."},
            {"title": "4. Final Onsite Loop (4-5 rounds)", "desc": "Rounds focusing heavily (80%) on Amazon's 16 Leadership Principles using STAR stories, with remaining focus on analytical design/execution cases."}
        ],
        "tips": [
            "Prepare 10-12 STAR stories mapped directly to Amazon's Leadership Principles, as every interviewer evaluates against them.",
            "Understand Amazon's document-driven culture. Focus on how you write specifications, PRFAQs, or narrative memos.",
            "Be highly quantitative: structure answers with data, metrics, and clear trade-off calculations."
        ]
    },
    "nvidia": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Initial screening on technical background, career goals, and compensation expectations."},
            {"title": "2. Hiring Manager Screen", "desc": "Deep technical and domain-specific discussion with the PM lead focusing on past projects and team-specific tech stack."},
            {"title": "3. Final Onsite Loop (3-6 rounds)", "desc": "Conversational rounds with PMs and engineering stakeholders focusing on real-world engineering trade-offs, developer ecosystems, and technical depth."}
        ],
        "tips": [
            "Possess a strong technical baseline. Understand GPU computing, developer tooling, parallel compute, or machine learning infrastructure.",
            "Highlight developer empathy. Nvidia PMs build products for developers, researchers, and system architects.",
            "Focus on behavioral history over generic design cases. Be ready to explain past project tradeoffs in detailed technical terms."
        ]
    },
    "meta": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "General background scan, interest in Meta, and leveling mapping."},
            {"title": "2. Product Sense Screen (45m)", "desc": "Initial case screening evaluating product design, user segmentations, feature prioritization, and success metrics."},
            {"title": "3. Onsite Loop (4-5 rounds)", "desc": "Rounds covering Product Sense (user-first design), Analytical Thinking (execution metrics, bug triage, trade-offs), and Leadership & Drive (behavioral alignment)."}
        ],
        "tips": [
            "For Analytical Thinking (Execution), focus heavily on data: practice metric frameworks, debugging, and counter-metric trade-offs.",
            "Structure all product sense cases: identify target user segments, define specific pain points, prioritize features, and define success metrics.",
            "Align recommendations with Meta's strategic pillars (e.g. connecting people, creator ecosystem, open-source AI, VR/AR)."
        ]
    },
    "tesla": {
        "stages": [
            {"title": "1. Recruiter Screen (20-60m)", "desc": "Alignment on logistics, interest in Tesla's mission, and general background check."},
            {"title": "2. Hiring Manager Screen", "desc": "Domain discussion evaluating hardware-software integration, manufacturing constraints, autonomy software, and team-specific fit."},
            {"title": "3. Onsite Loop & Presentation (5-6 rounds)", "desc": "Technical rounds with HW/SW engineering and PM leads, featuring a 45-minute technical presentation on a past complex project."}
        ],
        "tips": [
            "Demonstrate a high tolerance for ambiguity, speed, and shifting priorities. Tesla values quick decision cycles.",
            "Show extreme alignment with Tesla's mission: accelerating the transition to sustainable energy.",
            "Be highly technical: understand physical architectures, code optimization, or machine learning pipelines (e.g. FSD networks)."
        ]
    },
    "netflix": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Core screening checking alignment with Netflix's unique culture and values."},
            {"title": "2. Hiring Manager Screen", "desc": "Product discussion evaluating product sense, technical intuition, and initial culture fit."},
            {"title": "3. Culture Fit Screen (45m)", "desc": "Structured conversation with an HR Business Partner focusing on the Netflix Culture Memo values."},
            {"title": "4. Onsite Loop (4-5 rounds)", "desc": "Rounds with PMs, engineering leads, and design heads. Tests collaboration, conflict resolution, A/B testing standards, and metrics."}
        ],
        "tips": [
            "Memorize and internalize the Netflix Culture Memo. Expect interviewers to evaluate you against 'Freedom and Responsibility' values.",
            "Show strong familiarity with data-driven decision-making, specifically A/B testing, cohort metrics, and product instrumentation.",
            "Demonstrate absolute autonomy. Netflix values PMs who operate independently as high-performing 'stunners' without hand-holding."
        ]
    },
    "openai": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Initial sync on background, interest in AI, and alignment with safety mission."},
            {"title": "2. Hiring Manager Screen (1-2 rounds)", "desc": "Conversations focusing on products you shipped, technical capabilities, and strategic thinking."},
            {"title": "3. Take-Home / Case Assessment", "desc": "Skills-based assessment requiring candidates to develop a growth strategy or turn a vague prompt into a product spec deck."},
            {"title": "4. Onsite Loop (4-6 rounds)", "desc": "Highly ambiguous rounds covering Product Sense (vague single-sentence prompts), Execution, GTM Strategy, engineering collaboration, and safety/alignment."}
        ],
        "tips": [
            "Practice handling high ambiguity. Prepare to impose structure on abstract, single-sentence prompts with minimal guidance.",
            "Have deep AI/ML literacy: understand transformer architectures, context windows, tokens, latency, fine-tuning, and APIs.",
            "Highlight safety and ethical deployment. OpenAI evaluates candidates on their commitment to responsible AI development."
        ]
    },
    "anthropic": {
        "stages": [
            {"title": "1. Recruiter Screen (15-45m)", "desc": "Logistics check and screening on motivation, AI safety interest, and alignment with Anthropic's research philosophy."},
            {"title": "2. Hiring Manager Screen (45-60m)", "desc": "Discussion evaluating product judgment, cross-functional collaboration, and AI/ML domain understanding."},
            {"title": "3. Take-Home / Case Study", "desc": "Practical specification task requiring candidates to draft a spec for an AI feature, balancing UX with safety boundaries."},
            {"title": "4. Final Onsite Loop (3-5 rounds)", "desc": "Rounds focusing on Product Sense, Execution, and a dedicated, challenging Constitutional AI ethics and safety alignment round."}
        ],
        "tips": [
            "Study Anthropic's research papers (Constitutional AI, Model Alignment) to understand their safety-first product decisions.",
            "Treat the safety and ethics round as a hard technical interview; practice reasoning through complex AI deployment tradeoffs.",
            "Demonstrate developer empathy and familiarity with Claude, API developer tools, and model latency constraints."
        ]
    },
    "huggingface": {
        "stages": [
            {"title": "1. Recruiter/Hiring Manager Screen", "desc": "Conversational screen on background, open-source alignment, and ML familiarity."},
            {"title": "2. Conversations (2-3 rounds)", "desc": "Interviews with engineers and product leads exploring collaborative problem-solving and open-source contribution patterns."},
            {"title": "3. Role-Specific Take-Home/Assessment", "desc": "Practical exercise designing a hub integration, Spaces monetization, or API packaging strategy."},
            {"title": "4. Final Discussions & Debrief", "desc": "Final round with product leadership focusing on open-source community value vs commercial SaaS monetization."}
        ],
        "tips": [
            "Demonstrate a deep commitment to the open-source community and transparent development practices.",
            "Be active in or highly familiar with the Hugging Face Hub (models, datasets, Spaces, model hubs).",
            "Be ready to explain how to balance developer-centric open-source goodwill with enterprise monetization models."
        ]
    },
    "datadog": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Technical screen focusing on DevOps, cloud architectures, and observability landscape."},
            {"title": "2. Hiring Manager Screen (60m)", "desc": "Observability/telemetry case study focusing on developer personas and system dashboard designs."},
            {"title": "3. Onsite Panel (3-4 rounds)", "desc": "Rounds covering Observability System Design, Cloud APM/Logging Product Sense, Analytical Metrics (churn/adoption), and Behavioral alignment."}
        ],
        "tips": [
            "Understand highly technical developer personas (SREs, platform engineers, cloud architects) and their debugging workflows.",
            "Be fluent in observability concepts: APM, logging, metrics, tracing, distributed systems, and cloud-native Kubernetes environments.",
            "Focus on execution: showcase how you prioritize features and deliver value in short, iterative product cycles."
        ]
    },
    "cockroachlabs": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Initial scan checking database market experience and career alignment."},
            {"title": "2. Technical Work Experience Round", "desc": "Detailed walk-through of a complex product/feature you shipped, explaining architecture, trade-offs, and engineering negotiation."},
            {"title": "3. Product Positioning Exercise", "desc": "Exercise creating a B2B product messaging framework (elevator pitch, pillars, positioning) for a database feature."},
            {"title": "4. Onsite Panel (4-5 rounds)", "desc": "Exercise-based rounds evaluating developer experience (DX), distributed system trades, planning, and cultural values."}
        ],
        "tips": [
            "Review Cockroach Labs' open-sourced interview templates on GitHub to understand their exercise-based philosophy.",
            "Focus on B2B developer experience (DX) and technical scaling constraints (distributed consensus, latency, zero downtime).",
            "Be transparent and logical. They value clear problem-solving processes and structured thinking over guessing."
        ]
    },
    "disney": {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "General background alignment check, streaming/parks interest, and 'Why Disney?' scan."},
            {"title": "2. Hiring Manager Screen", "desc": "Discussion on PM methodologies, stakeholder management, and cross-functional shipping experience."},
            {"title": "3. Cross-Functional Panel (3-4 rounds)", "desc": "Interviews with engineering, design, and business leads evaluating collaboration, consumer product sense, and metrics (e.g. churn, CAC, watch-time)."},
            {"title": "4. Final/Director Round", "desc": "Conversations with senior directors evaluating long-term strategic vision, brand alignment, and leadership potential."}
        ],
        "tips": [
            "Articulate a clear, authentic passion for Disney's brand, storytelling, and user experiences across physical or streaming platforms.",
            "Know streaming metrics (CAC, churn, LTV, engagement, watch-time) and how technical features (player optimization, billing) affect them.",
            "Structure behavioral responses using the STAR method, emphasizing your ability to translate complex details for diverse creative/business stakeholders."
        ]
    }
}

def load_config_json():
    if not os.path.exists(CONFIG_PATH):
        return {"search_criteria": {}, "companies": []}
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_config_json(config_data):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config_data, f, indent=2)

def is_placeholder_url(url):
    if not url:
        return True
    url_lower = url.lower()
    placeholders = [
        "grounding-api-redirect", "search-results", "/search?", "/search/",
        "careers.google.com/jobs/results/?", "careers.microsoft.com",
        "jobs.apple.com", "openai.com/careers", "huggingface/jobs",
        "disneycareers.com", "cockroachlabs.com/careers/jobs",
        "/careers/jobs", "/careers", "/jobs"
    ]
    for p in placeholders:
        if p in url_lower:
            if "details/" in url_lower or "jobs/" in url_lower or "detail/" in url_lower or "job/" in url_lower or "pid=" in url_lower:
                continue
            return True
    return False

def get_insights(company_name):
    normalized = company_name.lower().replace(" ", "")
    for key, data in INTERVIEW_INSIGHTS.items():
        if key in normalized:
            return data
    return {
        "stages": [
            {"title": "1. Recruiter Screen (30m)", "desc": "Resume review, experience sanity check, and motivation for joining."},
            {"title": "2. Hiring Manager Screen (45m)", "desc": "Reviewing past PM successes/failures, alignment with company roadmap, and basic product design case."},
            {"title": "3. Product Case / Panel (60m)", "desc": "Product design or strategy exercise. Focus on user needs, prioritization, metrics, and technical viability."},
            {"title": "4. Onsite/Collaborative Rounds", "desc": "Interviews with Engineering leads, UI/UX designers, and cross-functional partners evaluating execution and communication."}
        ],
        "tips": [
            "Research the company's core product, business model, and recent news/releases.",
            "Prepare 3 strong STAR-method stories illustrating leadership, engineering collaboration, and data-driven prioritization.",
            "Be customer-centric: always start case questions by identifying who the target users are and what problems they face."
        ]
    }

def background_scan_thread():
    global is_scanning
    print("Background thread scan started...")
    try:
        python_executable = sys.executable
        script_path = os.path.join(SCRIPTS_DIR, "check_new_jobs.py")
        
        # Stream output in real-time using subprocess.Popen
        # Run python with -u flag to disable buffering of stdout/stderr
        process = subprocess.Popen(
            [python_executable, "-u", script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1
        )
        
        # Read stdout line by line and print to console
        while True:
            line = process.stdout.readline()
            if not line:
                break
            print(f"[Crawler] {line.rstrip()}")
            
        process.wait(timeout=900)  # Safe timeout of 15 minutes
        if process.returncode == 0:
            print("Background thread scan completed successfully.")
        else:
            print(f"Background thread scan failed with return code {process.returncode}")
    except subprocess.TimeoutExpired:
        print("Background thread scan timed out after 900 seconds.")
        if 'process' in locals():
            process.kill()
    except Exception as e:
        print("Error during background thread scan:", e)
        traceback.print_exc()
    finally:
        with scan_lock:
            is_scanning = False

@app.route('/')
def serve_index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory(app.static_folder, path)

def update_amazon_portal_url(url, selected_levels):
    import urllib.parse
    if not selected_levels:
        query = "product manager"
    else:
        if "Standard" in selected_levels or "Generic" in selected_levels:
            query = "product manager"
        elif "Senior" in selected_levels:
            query = "senior product manager"
        elif "Principal" in selected_levels:
            query = "principal product manager"
        elif "Staff" in selected_levels:
            query = "staff product manager"
        elif "Director" in selected_levels:
            query = "director of product"
        elif "Vice President" in selected_levels or "VP" in selected_levels:
            query = "vp of product"
        else:
            query = "product manager"
            
    parsed = urllib.parse.urlparse(url)
    params = urllib.parse.parse_qs(parsed.query)
    params["base_query"] = [query]
    new_query = urllib.parse.urlencode(params, doseq=True)
    return urllib.parse.urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))

# --- SEARCH CRITERIA API ---
@app.route('/api/criteria', methods=['GET'])
def get_criteria():
    config = load_config_json()
    return jsonify(config.get("search_criteria", {}))

@app.route('/api/criteria', methods=['POST'])
def save_criteria():
    try:
        data = request.get_json()
        config = load_config_json()
        
        # 1. Update search criteria
        if "search_criteria" in data:
            config["search_criteria"] = data["search_criteria"]
        else:
            config["search_criteria"] = {
                "locations": data.get("locations", []),
                "custom_keywords": data.get("custom_keywords", []),
                "strict_level_filtering": data.get("strict_level_filtering", False)
            }
            
        # 2. Update company levels if provided
        if "companies" in data:
            company_levels = {c.get("name", "").lower().strip(): c.get("levels", []) for c in data["companies"] if c.get("name")}
            for co in config.get("companies", []):
                co_name_lower = co.get("name", "").lower().strip()
                if co_name_lower in company_levels:
                    co["levels"] = company_levels[co_name_lower]
                    
        # 3. Update Amazon portal URL using Amazon's level settings
        amazon_levels = []
        for co in config.get("companies", []):
            if co.get("name", "").lower() == "amazon":
                amazon_levels = co.get("levels", [])
                co["portal_url"] = update_amazon_portal_url(co.get("portal_url", ""), amazon_levels)
                break
                
        save_config_json(config)
        return jsonify({"success": True, "message": "Search parameters and levels saved successfully."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# --- COMPANIES API ---
@app.route('/api/companies', methods=['GET'])
def get_companies():
    config = load_config_json()
    companies = config.get("companies", [])
    
    config_changed = False
    for co in companies:
        if "levels" not in co or co["levels"] is None:
            name_lower = co.get("name", "").lower().strip()
            if name_lower == "google":
                co["levels"] = ["Product Manager I", "Product Manager II", "Senior", "Group", "Director", "Senior Director", "Vice President"]
            elif name_lower == "meta":
                co["levels"] = ["Standard", "Leadership"]
            elif name_lower == "microsoft":
                co["levels"] = ["Standard", "Senior", "Principal", "Director", "Senior Director", "Vice President"]
            elif name_lower == "netflix":
                co["levels"] = ["Standard (up to L6)", "Group", "Director", "Senior Director", "Vice President"]
            elif name_lower in ["amazon", "nvidia"]:
                co["levels"] = ["Standard", "Senior", "Principal", "Director", "Senior Director", "Vice President"]
            elif name_lower in ["openai", "anthropic", "perplexity", "cursor"]:
                co["levels"] = ["Standard"]
            elif name_lower == "apple":
                co["levels"] = ["Standard", "Director", "Senior Director"]
            elif name_lower == "datadog":
                co["levels"] = ["Standard", "Senior", "Group", "Staff", "Principal", "Director", "Senior Director", "Vice President", "Chief Product Officer"]
            else:
                co["levels"] = ["Standard", "Senior", "Staff", "Principal", "Director", "Senior Director", "Vice President", "Chief Product Officer"]
            config_changed = True
            
    if config_changed:
        config["companies"] = companies
        save_config_json(config)
        
    return jsonify(companies)

def single_company_scan_thread(company_name):
    global is_scanning
    print(f"Background single-company scan started for {company_name}...")
    try:
        python_executable = sys.executable
        script_path = os.path.join(SCRIPTS_DIR, "check_new_jobs.py")
        
        process = subprocess.Popen(
            [python_executable, "-u", script_path, company_name],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1
        )
        
        while True:
            line = process.stdout.readline()
            if not line:
                break
            print(f"[SingleCrawler:{company_name}] {line.rstrip()}")
            
        process.wait(timeout=240)  # Double the timeout to 4 minutes
        if process.returncode == 0:
            print(f"Background single-company scan completed for {company_name}.")
        else:
            print(f"Background single-company scan failed for {company_name} with return code {process.returncode}")
    except subprocess.TimeoutExpired:
        print(f"Background single-company scan timed out for {company_name} after 240 seconds.")
        if 'process' in locals():
            process.kill()
    except Exception as e:
        print(f"Error during background single-company scan for {company_name}:", e)
        traceback.print_exc()
    finally:
        with scan_lock:
            is_scanning = False

def delete_company_jobs_from_excel(company_name):
    if not os.path.exists(TRACKER_PATH):
        return 0
    try:
        wb = openpyxl.load_workbook(TRACKER_PATH)
        if "Job Leads" in wb.sheetnames:
            ws = wb["Job Leads"]
            deleted_count = 0
            # Iterate backwards to avoid row offset shift bugs
            for r_idx in range(ws.max_row, 1, -1):
                company_val = ws.cell(row=r_idx, column=3).value
                if company_val and company_val.lower().strip() == company_name.lower().strip():
                    ws.delete_rows(r_idx)
                    deleted_count += 1
            if deleted_count > 0:
                wb.save(TRACKER_PATH)
            return deleted_count
    except Exception as e:
        print(f"Error deleting jobs for company '{company_name}' from Excel:", e)
    return 0

@app.route('/api/companies', methods=['POST'])
def add_company():
    try:
        new_co = request.get_json()
        name = new_co.get("name", "").strip()
        if not name:
            return jsonify({"error": "Company name is required"}), 400
            
        config = load_config_json()
        companies = config.get("companies", [])
        if any(c.get("name", "").lower() == name.lower() for c in companies):
            return jsonify({"error": f"Company '{name}' is already in your tracked list."}), 400
            
        portal_url = new_co.get("portal_url", "").strip()
        if not portal_url:
            return jsonify({"error": "Careers Search Portal URL is required"}), 400
            
        # Detect platform and board ID from careers portal URL
        import urllib.parse
        platform = "ddg"
        board_id = ""
        portal_url_lower = portal_url.lower()
        name_lower = name.lower()

        if "doordash" in portal_url_lower or name_lower == "doordash":
            platform = "greenhouse"
            board_id = "doordashusa"
        elif "google.com" in portal_url_lower:
            platform = "google"
        elif "microsoft.com" in portal_url_lower:
            platform = "microsoft"
        elif "netflix.com" in portal_url_lower:
            platform = "netflix"
        elif "amazon.jobs" in portal_url_lower or "amazon.com" in portal_url_lower:
            platform = "amazon"
        elif "nvidia.com" in portal_url_lower:
            platform = "nvidia"
            board_id = "NVIDIAExternalCareerSite"
        elif "metacareers.com" in portal_url_lower or "meta.com" in portal_url_lower:
            platform = "meta"
        elif "tesla.com" in portal_url_lower:
            platform = "tesla"
        elif "apple.com" in portal_url_lower:
            platform = "apple"
        elif "greenhouse.io" in portal_url_lower:
            platform = "greenhouse"
            parsed_url = urllib.parse.urlparse(portal_url)
            path_parts = [p for p in parsed_url.path.split('/') if p]
            if path_parts:
                board_id = path_parts[-1]
        elif "ashbyhq.com" in portal_url_lower:
            platform = "ashby"
            parsed_url = urllib.parse.urlparse(portal_url)
            path_parts = [p for p in parsed_url.path.split('/') if p]
            if path_parts:
                board_id = path_parts[-1]
        elif "workable.com" in portal_url_lower:
            platform = "workable"
            parsed_url = urllib.parse.urlparse(portal_url)
            path_parts = [p for p in parsed_url.path.split('/') if p]
            if path_parts:
                board_id = path_parts[0]
                
        capability = "portal_only" if platform in ["tesla", "ddg"] else "active_sync"

        # Determine default levels if none provided in POST request
        levels = new_co.get("levels", [])
        if not levels:
            name_lower = name.lower().strip()
            if name_lower == "google":
                levels = ["Product Manager I", "Product Manager II", "Senior", "Group", "Director", "Senior Director", "Vice President"]
            elif name_lower == "meta":
                levels = ["Standard", "Leadership"]
            elif name_lower == "microsoft":
                levels = ["Standard", "Senior", "Principal", "Director", "Senior Director", "Vice President"]
            elif name_lower == "netflix":
                levels = ["Standard (up to L6)", "Group", "Director", "Senior Director", "Vice President"]
            elif name_lower in ["amazon", "nvidia"]:
                levels = ["Standard", "Senior", "Principal", "Director", "Senior Director", "Vice President"]
            elif name_lower in ["openai", "anthropic", "perplexity", "cursor"]:
                levels = ["Standard"]
            elif name_lower == "apple":
                levels = ["Standard", "Director", "Senior Director"]
            elif name_lower == "datadog":
                levels = ["Standard", "Senior", "Group", "Staff", "Principal", "Director", "Senior Director", "Vice President", "Chief Product Officer"]
            else:
                levels = ["Standard", "Senior", "Staff", "Principal", "Director", "Senior Director", "Vice President", "Chief Product Officer"]

        companies.append({
            "name": name,
            "platform": platform,
            "board_id": board_id,
            "cohort": new_co.get("cohort", "Non-Mag 7").strip(),
            "hq": "N/A",
            "founded": "N/A",
            "revenue": "N/A",
            "employees": "N/A",
            "domain": "N/A",
            "capability": capability,
            "portal_url": portal_url,
            "levels": levels
        })
        config["companies"] = companies
        save_config_json(config)

        # Trigger single company scan immediately in background
        global is_scanning
        with scan_lock:
            is_scanning = True
        thread = threading.Thread(target=single_company_scan_thread, args=(name,))
        thread.daemon = True
        thread.start()

        return jsonify({"success": True, "message": f"Company '{name}' added. Starting immediate crawl scan in the background..."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/companies/<string:company_name>', methods=['DELETE'])
def delete_company(company_name):
    try:
        config = load_config_json()
        companies = config.get("companies", [])
        
        company_exists = any(c.get("name", "").lower() == company_name.lower().strip() for c in companies)
        if not company_exists:
            return jsonify({"error": f"Company '{company_name}' not found."}), 404
            
        # Filter out company
        updated_companies = [c for c in companies if c.get("name", "").lower() != company_name.lower().strip()]
        config["companies"] = updated_companies
        save_config_json(config)
        
        # Delete jobs related to that company from Excel
        deleted_jobs = delete_company_jobs_from_excel(company_name)
        
        # Invalidate in-memory cache to force immediate reload
        JOBS_CACHE["last_loaded_mtime"] = 0
        
        return jsonify({
            "success": True, 
            "message": f"Company '{company_name}' removed. Deleted {deleted_jobs} related job listings from spreadsheet."
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- COMPANY SPECIFIC INFO ---
@app.route('/api/company-info', methods=['GET'])
def get_company_info():
    company_name = request.args.get("company", "").strip().lower()
    config = load_config_json()
    
    for co in config.get("companies", []):
        if co.get("name", "").lower() == company_name:
            return jsonify(co)
            
    return jsonify({
        "name": request.args.get("company", "Generic"),
        "hq": "Information not tracked",
        "founded": "N/A",
        "revenue": "Unknown / Private",
        "employees": "N/A",
        "domain": "Technology"
    })

# --- JOBS LIST API (ASYNCHRONOUS AUTO-TRIGGER GENTLE FIRST SCAN) ---
@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    try:
        global auto_scan_triggered, is_scanning
        
        # Trigger background scan automatically on first load in session
        if not auto_scan_triggered:
            with scan_lock:
                if not is_scanning:
                    is_scanning = True
                    auto_scan_triggered = True
                    thread = threading.Thread(target=background_scan_thread)
                    thread.daemon = True
                    thread.start()
                    print("Auto-first scan triggered asynchronously in background thread.")
        
        jobs, last_scanned = load_jobs_from_excel()
        
        return jsonify({
            "jobs": jobs,
            "last_scanned": last_scanned,
            "is_scanning": is_scanning
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:row_id>/status', methods=['POST'])
def update_job_status(row_id):
    try:
        data = request.get_json()
        new_status = data.get("status")
        
        if new_status not in ["Lead", "Consideration", "Archived"]:
            return jsonify({"error": f"Invalid status: {new_status}"}), 400
            
        if not os.path.exists(TRACKER_PATH):
            return jsonify({"error": "Job leads tracker file not found."}), 404
            
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb["Job Leads"]
        
        if row_id < 2 or row_id > ws.max_row:
            return jsonify({"error": f"Invalid job ID: {row_id}"}), 400
            
        ws.cell(row=row_id, column=7).value = new_status
        
        if new_status == "Consideration":
            ws.cell(row=row_id, column=1).value = "[x]"
            ws.cell(row=row_id, column=15).value = ""
        elif new_status == "Archived":
            ws.cell(row=row_id, column=1).value = "[ ]"
            # Set to User Archived unless it's already set (e.g. was auto-archived as Closed)
            if not ws.cell(row=row_id, column=15).value:
                ws.cell(row=row_id, column=15).value = "User Archived"
        else:
            ws.cell(row=row_id, column=1).value = "[ ]"
            ws.cell(row=row_id, column=15).value = ""
            
        wb.save(TRACKER_PATH)
        
        # Invalidate in-memory cache to force immediate reload
        JOBS_CACHE["last_loaded_mtime"] = 0
        
        return jsonify({"success": True, "message": f"Job row {row_id} updated to {new_status}."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:row_id>', methods=['DELETE'])
def delete_job(row_id):
    try:
        if not os.path.exists(TRACKER_PATH):
            return jsonify({"error": "Job leads tracker file not found."}), 404
            
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb["Job Leads"]
        
        if row_id < 2 or row_id > ws.max_row:
            return jsonify({"error": f"Invalid job ID: {row_id}"}), 400
            
        ws.delete_rows(row_id)
        wb.save(TRACKER_PATH)
        
        # Invalidate in-memory cache to force immediate reload
        JOBS_CACHE["last_loaded_mtime"] = 0
        
        return jsonify({"success": True, "message": f"Job row {row_id} deleted successfully."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# --- JOBS TRACKING DETAILS API ---
@app.route('/api/jobs/<int:row_id>/tracking', methods=['POST'])
def update_job_tracking(row_id):
    try:
        data = request.get_json()
        app_status = data.get("app_status", "Not Applied")
        app_outcome = data.get("app_outcome", "Active / Pending")
        resume_url = data.get("resume_url", "")
        
        if not os.path.exists(TRACKER_PATH):
            return jsonify({"error": "Job leads tracker file not found."}), 404
            
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb["Job Leads"]
        
        if row_id < 2 or row_id > ws.max_row:
            return jsonify({"error": f"Invalid job ID: {row_id}"}), 400
            
        # Ensure tracking headers exist in Excel
        if ws.max_column < 14 or ws.cell(row=1, column=12).value is None:
            ws.cell(row=1, column=11).value = "Compensation"
            ws.cell(row=1, column=12).value = "App Status"
            ws.cell(row=1, column=13).value = "App Outcome"
            ws.cell(row=1, column=14).value = "Resume Link"
            
        ws.cell(row=row_id, column=12).value = app_status
        ws.cell(row=row_id, column=13).value = app_outcome
        ws.cell(row=row_id, column=14).value = resume_url
        
        wb.save(TRACKER_PATH)
        
        # Invalidate in-memory cache to force immediate reload
        JOBS_CACHE["last_loaded_mtime"] = 0
        
        return jsonify({"success": True, "message": "Application tracking updated successfully in Excel."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# --- ASYNC SCAN TRIGGER API ---
@app.route('/api/scan', methods=['POST'])
def run_scan():
    global is_scanning
    with scan_lock:
        if is_scanning:
            return jsonify({"success": False, "error": "A scan is already in progress in the background."}), 400
        is_scanning = True
        
    thread = threading.Thread(target=background_scan_thread)
    thread.daemon = True
    thread.start()
    
    return jsonify({
        "success": True,
        "message": "Crawl scan started asynchronously in the background thread."
    })

# --- SCAN STATUS API ---
@app.route('/api/scan-status', methods=['GET'])
def get_scan_status():
    global is_scanning
    return jsonify({"is_scanning": is_scanning})

@app.route('/api/job-description', methods=['GET'])
def get_job_description():
    url = request.args.get("url")
    if not url or is_placeholder_url(url):
        return jsonify({
            "description": "<p><strong>Placeholder URL:</strong> No direct job description can be parsed. This occurs when the automated crawl only identifies the careers landing portal. Please follow the link below to search for this role on the company's careers site.</p>"
        })
        
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5"
        }
        
        r = requests.get(url, headers=headers, timeout=12)
        if r.status_code != 200:
            return jsonify({
                "description": f"<p><strong>Failed to load content:</strong> The remote server returned status code {r.status_code}. <a href='{url}' target='_blank' class='inline-link'>Click here to open the page directly.</a></p>"
            })
            
        soup = BeautifulSoup(r.text, 'html.parser')
        
        desc_html = ""
        
        # Try to parse Apple Careers first
        if "jobs.apple.com" in url.lower():
            import re
            match = re.search(r'window\.__staticRouterHydrationData\s*=\s*JSON\.parse\("((?:[^"\\]|\\.)*)"\);', r.text)
            if match:
                try:
                    escaped_json = match.group(1)
                    raw_json_str = json.loads(f'"{escaped_json}"')
                    data = json.loads(raw_json_str)
                    job_data = data.get("loaderData", {}).get("jobDetails", {}).get("jobsData", {})
                    if job_data:
                        summary = job_data.get('jobSummary', '')
                        if summary:
                            desc_html += f"<h3>Job Summary</h3><p>{summary}</p>"
                        
                        desc_text = job_data.get('description', '')
                        if desc_text:
                            desc_text_html = desc_text.replace('\n', '<br/>')
                            desc_html += f"<h3>Description</h3><p>{desc_text_html}</p>"
                            
                        min_quals = job_data.get('minimumQualifications', '')
                        if min_quals:
                            quals_list = "".join(f"<li>{q.strip()}</li>" for q in min_quals.split("\n") if q.strip())
                            desc_html += f"<h3>Minimum Qualifications</h3><ul>{quals_list}</ul>"
                            
                        pref_quals = job_data.get('preferredQualifications', '')
                        if pref_quals:
                            pref_list = "".join(f"<li>{q.strip()}</li>" for q in pref_quals.split("\n") if q.strip())
                            desc_html += f"<h3>Preferred Qualifications</h3><ul>{pref_list}</ul>"
                except Exception as e:
                    print("Error parsing Apple details page hydration data:", e)
                    
        # 1. Try to extract description from JSON-LD schema (JobPosting)
        if not desc_html:
            for s in soup.find_all('script', type='application/ld+json'):
                try:
                    data = json.loads(s.string or "")
                    if isinstance(data, dict):
                        if data.get("@type") == "JobPosting" and data.get("description"):
                            desc_html = data["description"]
                            break
                    elif isinstance(data, list):
                        for item in data:
                            if isinstance(item, dict) and item.get("@type") == "JobPosting" and item.get("description"):
                                desc_html = item["description"]
                                break
                        if desc_html:
                            break
                except Exception:
                    pass
                
        # If not found via JSON-LD, fall back to parsing the HTML body
        if not desc_html:
            for element in soup(["script", "style", "nav", "footer", "header", "noscript"]):
                element.decompose()
        url_lower = url.lower()
        
        if "greenhouse.io" in url_lower:
            container = soup.find(id="content") or soup.find(class_="job-body") or soup.find(id="main")
            if container:
                desc_html = str(container)
        elif "ashbyhq.com" in url_lower:
            container = soup.find(class_="_jobPosting_") or soup.find(class_="job-posting") or soup.find(id="app")
            if container:
                desc_html = str(container)
        elif "workable.com" in url_lower:
            container = soup.find(attrs={"data-ui": "job-description"}) or soup.find(class_="section--job-description")
            if container:
                desc_html = str(container)
        elif "careers.google.com" in url_lower or "google.com/about/careers" in url_lower:
            container = soup.find(class_="job-description") or soup.find(attrs={"itemprop": "description"}) or soup.find(class_="gc-card__job-description")
            if container:
                desc_html = str(container)
                
        if not desc_html:
            candidates = soup.find_all(lambda tag: tag.name in ['div', 'section', 'article'] and 
                                      (any('description' in str(attr).lower() for attr in tag.attrs.values()) or
                                       any('job-body' in str(attr).lower() for attr in tag.attrs.values()) or
                                       any('posting' in str(attr).lower() for attr in tag.attrs.values())))
            
            if candidates:
                candidates = sorted(candidates, key=lambda tag: len(tag.text), reverse=True)
                desc_html = str(candidates[0])
            else:
                body = soup.find('body')
                if body:
                    main = body.find('main') or body.find('article') or body.find(id='content')
                    if main:
                        desc_html = str(main)
                    else:
                        paragraphs = body.find_all(['p', 'ul', 'ol', 'h1', 'h2', 'h3'])
                        desc_html = "".join(str(p) for p in paragraphs[:30])
                        
        if not desc_html:
            desc_html = f"<p>Could not extract the description text automatically. Please <a href='{url}' target='_blank' class='inline-link'>view the job posting on the careers page.</a></p>"
            
        return jsonify({
            "description": desc_html
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({
            "description": f"<p><strong>Error loading description:</strong> {str(e)}. <a href='{url}' target='_blank' class='inline-link'>Open job listing directly.</a></p>"
        })

@app.route('/api/insights', methods=['GET'])
def get_company_insights():
    company = request.args.get("company", "")
    insights = get_insights(company)
    return jsonify(insights)

# --- LINKEDIN CONNECTIONS API ---
CONNECTIONS_PATH = r"G:\My Drive\.Agents\1_Projects\2026 Job Search\board\connections.json"

def load_connections():
    if not os.path.exists(CONNECTIONS_PATH):
        return []
    try:
        with open(CONNECTIONS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print("Error loading connections:", e)
        return []

def save_connections(data):
    try:
        with open(CONNECTIONS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print("Error saving connections:", e)
        return False

@app.route('/api/connections', methods=['GET'])
def get_connections():
    return jsonify(load_connections())

@app.route('/api/connections/clear', methods=['POST'])
def clear_connections():
    if save_connections([]):
        return jsonify({"success": True, "message": "Connections cleared successfully."})
    return jsonify({"error": "Failed to clear connections."}), 500

@app.route('/api/connections/upload', methods=['POST'])
def upload_connections():
    import csv
    import io
    import urllib.parse
    
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request."}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file."}), 400
        
    if not file.filename.endswith('.csv'):
        return jsonify({"error": "Only CSV files are supported."}), 400
        
    try:
        # Read the file content as text
        stream = io.StringIO(file.stream.read().decode("utf-8"), newline=None)
        csv_content = stream.read()
        
        lines = csv_content.splitlines()
        header_idx = -1
        for idx, line in enumerate(lines):
            # Check if this line looks like the header
            if "First Name" in line and "Last Name" in line:
                header_idx = idx
                break
                
        if header_idx == -1:
            header_idx = 0
            
        csv_data = "\n".join(lines[header_idx:])
        reader = csv.DictReader(io.StringIO(csv_data))
        
        # Normalize header names (strip whitespace and remove BOM if present)
        if reader.fieldnames:
            reader.fieldnames = [name.strip().replace('\ufeff', '') for name in reader.fieldnames]
            
        parsed_connections = []
        for row in reader:
            first_name = (row.get("First Name") or "").strip()
            last_name = (row.get("Last Name") or "").strip()
            company = (row.get("Company") or "").strip()
            position = (row.get("Position") or "").strip()
            url = (row.get("URL") or "").strip()
            
            # Skip rows where name is empty
            if not first_name and not last_name:
                continue
                
            # If no URL is provided, generate a search URL
            if not url:
                query = f"{first_name} {last_name}".strip()
                url = f"https://www.linkedin.com/search/results/people/?keywords={urllib.parse.quote(query)}"
                
            parsed_connections.append({
                "first_name": first_name,
                "last_name": last_name,
                "company": company,
                "position": position,
                "url": url
            })
            
        mode = request.form.get('mode', 'replace')
        if mode == 'append':
            existing = load_connections()
            
            # Helper to generate unique signature for deduplication
            def get_signature(c):
                url_clean = c.get('url', '').strip().lower()
                if url_clean and "search/results" not in url_clean:
                    return f"url:{url_clean}"
                name_clean = f"{c.get('first_name','')}_{c.get('last_name','')}".strip().lower().replace(" ","")
                comp_clean = c.get('company','').strip().lower().replace(" ","")
                return f"info:{name_clean}@{comp_clean}"
                
            existing_sigs = {get_signature(x) for x in existing}
            
            added_count = 0
            for new_conn in parsed_connections:
                sig = get_signature(new_conn)
                if sig not in existing_sigs:
                    existing.append(new_conn)
                    existing_sigs.add(sig)
                    added_count += 1
            
            save_connections(existing)
            message = f"Successfully appended {added_count} new connections (skipped {len(parsed_connections) - added_count} duplicates)."
            result_connections = existing
        else:
            save_connections(parsed_connections)
            message = f"Successfully replaced database with {len(parsed_connections)} connections."
            result_connections = parsed_connections
            
        return jsonify({
            "success": True, 
            "message": message,
            "connections": result_connections
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to parse CSV file: {str(e)}"}), 500

if __name__ == '__main__':
    print("Starting My Job Board backend on http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)
