import socket
# Force IPv4 only to prevent connection hangs on misconfigured IPv6 environments
old_getaddrinfo = socket.getaddrinfo
def new_getaddrinfo(*args, **kwargs):
    responses = old_getaddrinfo(*args, **kwargs)
    return [r for r in responses if r[0] == socket.AF_INET]
socket.getaddrinfo = new_getaddrinfo

import os
import sys
import requests
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import re
import urllib.parse
from bs4 import BeautifulSoup
import urllib3

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

CONFIG_PATH = r"G:\My Drive\.Agents\1_Projects\2026 Job Search\board\config.json"
TRACKER_PATH = r"G:\My Drive\.Agents\1_Projects\2026 Job Search\Job_Leads_Tracker.xlsx"

def load_config():
    if not os.path.exists(CONFIG_PATH):
        # Fallback default
        return {
            "search_criteria": {
                "locations": ["New York", "NYC", "Remote"],
                "role_levels": ["Senior", "Principal", "Staff", "Director", "VP"],
                "custom_keywords": ["product manager", "product management", "technical product manager"],
                "strict_level_filtering": False
            },
            "companies": []
        }
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
        
    # Auto-backfill default levels if missing
    config_changed = False
    for co in config.get("companies", []):
        if "levels" not in co or co["levels"] is None:
            name_lower = co.get("name", "").lower().strip()
            if name_lower == "google":
                co["levels"] = ["Product Manager I", "Product Manager II", "Senior", "Group", "Director", "Senior Director", "Vice President"]
            elif name_lower == "meta":
                co["levels"] = ["Standard", "Leadership"]
            elif name_lower == "microsoft":
                co["levels"] = ["Standard", "Senior", "Principal", "Director", "Senior Director", "Vice President"]
            elif name_lower == "netflix":
                co["levels"] = ["Standard (up to L6)", "Group", "Director", "Senior Director", "Vice President"]
            elif name_lower in ["amazon", "nvidia"]:
                co["levels"] = ["Standard", "Senior", "Principal", "Director", "Senior Director", "Vice President"]
            elif name_lower in ["openai", "anthropic", "perplexity", "cursor"]:
                co["levels"] = ["Standard"]
            elif name_lower == "apple":
                co["levels"] = ["Standard", "Director", "Senior Director"]
            elif name_lower == "datadog":
                co["levels"] = ["Standard", "Senior", "Group", "Staff", "Principal", "Director", "Senior Director", "Vice President", "Chief Product Officer"]
            else:
                co["levels"] = ["Standard", "Senior", "Staff", "Principal", "Director", "Senior Director", "Vice President", "Chief Product Officer"]
            config_changed = True
            
    if config_changed:
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Error saving auto-backfilled config in crawler: {e}")
            
    return config

def normalize_string(s):
    if not s:
        return ""
    # Retain parentheses because they distinguish specific focus areas (e.g. Netflix PM focus teams)
    s = s.replace("-", " ").replace(",", " ")
    return " ".join(s.split()).lower().strip()

def is_placeholder_url(url):
    if not url:
        return True
    url_lower = url.lower()
    placeholders = [
        "grounding-api-redirect", "search-results", "/search?", "/search/",
        "careers.google.com/jobs/results/?", "careers.microsoft.com",
        "jobs.apple.com", "openai.com/careers", "huggingface/jobs",
        "disneycareers.com", "cockroachlabs.com/careers/jobs",
        "/careers/jobs", "/careers", "/jobs"
    ]
    for p in placeholders:
        if p in url_lower:
            if "details/" in url_lower or "jobs/" in url_lower or "detail/" in url_lower or "job/" in url_lower or "pid=" in url_lower:
                continue
            return True
    return False


def is_job_url_closed(url, company_name):
    if not url or is_placeholder_url(url):
        return False
        
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5"
    }
    
    try:
        # 1. Greenhouse
        if "greenhouse.io" in url.lower():
            r = requests.get(url, headers=headers, timeout=10, allow_redirects=True, verify=False)
            if r.status_code == 404:
                return True
            if "/jobs/" not in r.url.lower() and "/postings/" not in r.url.lower():
                return True
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text().lower()
            if "job is no longer available" in text or "no longer open" in text or "position you are looking for has been filled" in text:
                return True
            return False
            
        # 2. Ashby
        elif "ashbyhq.com" in url.lower():
            r = requests.get(url, headers=headers, timeout=10, allow_redirects=True, verify=False)
            if r.status_code == 404:
                return True
            if "posting" not in r.url.lower():
                return True
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text().lower()
            if "no longer available" in text or "posting is closed" in text or "position you are looking for has been filled" in text:
                return True
            return False
            
        # 3. Workable
        elif "workable.com" in url.lower():
            r = requests.get(url, headers=headers, timeout=10, allow_redirects=True, verify=False)
            if r.status_code == 404:
                return True
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text().lower()
            if "no longer available" in text or "job is closed" in text or "not accepting applications" in text:
                return True
            return False
            
        # 4. Microsoft
        elif "careers.microsoft.com" in url.lower() or "apply.careers.microsoft.com" in url.lower():
            pid_match = re.search(r'/job/(\d+)', url)
            if not pid_match:
                pid_match = re.search(r'position_id=(\d+)', url)
            if pid_match:
                pid = pid_match.group(1)
                detail_url = f"https://apply.careers.microsoft.com/api/pcsx/position_details?domain=microsoft.com&position_id={pid}"
                r = requests.get(detail_url, headers=headers, timeout=10, verify=False)
                if r.status_code == 200:
                    try:
                        data = r.json()
                        if not data.get('data') or not data.get('data', {}).get('jobDescription'):
                            return True
                    except:
                        return True
                else:
                    return True
            return False
            
        # 5. Netflix
        elif "explore.jobs.netflix.net" in url.lower():
            parsed = urllib.parse.urlparse(url)
            query_params = urllib.parse.parse_qs(parsed.query)
            pid_list = query_params.get("pid")
            if pid_list:
                pid = pid_list[0]
                r = requests.get(url, headers=headers, timeout=10, verify=False)
                if r.status_code == 200:
                    soup = BeautifulSoup(r.text, 'html.parser')
                    code_tag = soup.find('code', id='smartApplyData')
                    if code_tag:
                        try:
                            config = json.loads(code_tag.get_text())
                            if not config.get("positions"):
                                return True
                        except:
                            return True
                    else:
                        text = soup.get_text().lower()
                        if "no longer available" in text or "not found" in text:
                            return True
                else:
                    return True
            return False
            
        # 6. Apple
        elif "jobs.apple.com" in url.lower():
            r = requests.get(url, headers=headers, timeout=10, verify=False)
            if r.status_code == 404:
                return True
            if r.status_code in [400, 403]:
                return False
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text().lower()
            if "no longer available" in text or "page not found" in text or "invalid job" in text:
                return True
            return False
            
        # 7. Amazon
        elif "amazon.jobs" in url.lower():
            r = requests.get(url, headers=headers, timeout=10, allow_redirects=True, verify=False)
            if r.status_code == 404:
                return True
            if r.status_code in [400, 403]:
                return False
            if "jobs" not in r.url.lower():
                return True
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text().lower()
            indicators = ["no longer available", "position has been filled", "job is closed", "not accepting applications", "posting has expired", "job is no longer active"]
            if any(ind in text for ind in indicators):
                return True
            return False
            
        # 8. Meta
        elif "metacareers.com" in url.lower():
            r = requests.get(url, headers=headers, timeout=10, allow_redirects=True, verify=False)
            if r.status_code == 404:
                return True
            if r.status_code in [400, 403]:
                return False
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text().lower()
            if "no longer available" in text or "job you are looking for is closed" in text:
                return True
            return False
            
        # 9. Generic
        else:
            r = requests.get(url, headers=headers, timeout=10, allow_redirects=True, verify=False)
            if r.status_code == 404:
                return True
            if r.status_code in [400, 403]:
                return False
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text().lower()
            indicators = [
                "no longer available",
                "position is no longer available",
                "job is no longer available",
                "this job is closed",
                "no longer accepting applications",
                "job posting has expired",
                "listing has expired",
                "position has been filled",
                "page you are looking for has been filled",
                "invalid job"
            ]
            if any(ind in text for ind in indicators):
                return True
            return False
            
    except Exception as e:
        print(f"Error checking status for URL {url}: {e}")
        return False

def extract_salary(text):
    if not text:
        return None
    
    import html
    unescaped = html.unescape(text)
    soup = BeautifulSoup(unescaped, 'html.parser')
    clean_text = soup.get_text(separator=' ')
    clean_text = re.sub(r'\s+', ' ', clean_text)
    
    pattern = r'(\$?)\s*([0-9]{2,3}(?:,[0-9]{3})*(?:\.[0-9]{2})?|[0-9]{5,6})\s*([kK]?)\s*(USD|/yr|/year|/hr|/hour|per year|per hour)?\s*(?:-|to|and|through|\u2013|\u2014)\s*(\$?)\s*([0-9]{2,3}(?:,[0-9]{3})*(?:\.[0-9]{2})?|[0-9]{5,6})\s*([kK]?)\s*(USD|/yr|/year|/hr|/hour|per year|per hour)?'
    
    matches = re.finditer(pattern, clean_text)
    for m in matches:
        has_dollar = bool(m.group(1) or m.group(5))
        val1_str = m.group(2)
        k1 = m.group(3)
        unit1 = m.group(4) or ""
        val2_str = m.group(6)
        k2 = m.group(7)
        unit2 = m.group(8) or ""
        
        # Clean commas and decimals
        v1_clean = re.sub(r'[,\.]', '', val1_str).strip()
        if '.' in val1_str:
            v1_clean = val1_str.split('.')[0].replace(',', '')
            
        v2_clean = re.sub(r'[,\.]', '', val2_str).strip()
        if '.' in val2_str:
            v2_clean = val2_str.split('.')[0].replace(',', '')
            
        try:
            v1 = int(v1_clean)
            v2 = int(v2_clean)
            
            # Apply k/K multipliers
            if k1.lower() == 'k':
                v1 *= 1000
            if k2.lower() == 'k':
                v2 *= 1000
                
            if v1 < 1000 and (k1.lower() == 'k' or k2.lower() == 'k' or v2 >= 30000):
                v1 *= 1000
            if v2 < 1000 and (k1.lower() == 'k' or k2.lower() == 'k' or v1 >= 30000):
                v2 *= 1000
                
            # Range check for annual salary
            if (30000 <= v1 <= 1000000 and 30000 <= v2 <= 1000000):
                if v1 > v2:
                    v1, v2 = v2, v1
                return f"${v1:,} - ${v2:,}"
                
            # Range check for hourly rate (e.g., 30 to 300)
            elif (30 <= v1 <= 300 and 30 <= v2 <= 300):
                unit_str = (unit1 + unit2).lower()
                is_hourly = 'hr' in unit_str or 'hour' in unit_str or 'hourly' in unit_str
                
                if has_dollar or is_hourly:
                    if v1 > v2:
                        v1, v2 = v2, v1
                    return f"${v1}/hr - ${v2}/hr"
        except:
            pass
    return None

# DuckDuckGo HTML Search Helper
def search_ddg(query):
    url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(query)}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    for attempt in range(3):
        try:
            r = requests.get(url, headers=headers, timeout=10)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                links = []
                for a in soup.find_all('a', class_='result__snippet', href=True):
                    href = a['href']
                    if "/l/?uddg=" in href:
                        parsed = urllib.parse.urlparse(href)
                        query_params = urllib.parse.parse_qs(parsed.query)
                        real_url = query_params.get("uddg", [None])[0]
                        if real_url:
                            links.append(real_url)
                    elif href.startswith("http"):
                        links.append(href)
                for a in soup.find_all('a', class_='result__a', href=True):
                    href = a['href']
                    if "/l/?uddg=" in href:
                        parsed = urllib.parse.urlparse(href)
                        query_params = urllib.parse.parse_qs(parsed.query)
                        real_url = query_params.get("uddg", [None])[0]
                        if real_url:
                            links.append(real_url)
                    elif href.startswith("http"):
                        links.append(href)
                if links:
                    return list(set(links))
            elif r.status_code == 202:
                # Meta refresh redirect
                soup = BeautifulSoup(r.text, 'html.parser')
                meta_refresh = soup.find('meta', attrs={'http-equiv': 'refresh'})
                if meta_refresh:
                    content = meta_refresh.get('content', '')
                    if 'url=' in content:
                        refresh_url = content.split('url=')[-1]
                        if refresh_url.startswith('/'):
                            refresh_url = f"https://html.duckduckgo.com{refresh_url}"
                        url = refresh_url
                        continue
        except Exception as e:
            print(f"DDG Search attempt {attempt} failed: {e}")
        import time
        time.sleep(1)
    return []

def title_matches_company_levels(title, company_name, config=None):
    if config is None:
        config = load_config()
    title_lower = title.lower()
    
    # Find company in config
    companies = config.get("companies", [])
    target_company = None
    for co in companies:
        if co.get("name", "").lower().strip() == company_name.lower().strip():
            target_company = co
            break
            
    if not target_company:
        return True
        
    selected_levels = target_company.get("levels")
    if selected_levels is None:
        return True
        
    if not selected_levels:
        return False
        
    selected_levels_lower = [l.lower().strip() for l in selected_levels]
    company_name_lower = company_name.lower().strip()
    
    # Helper to check if a title is a generic PM title, respecting abbreviations
    def is_generic_pm_title(t_lower):
        if "product manager" in t_lower or "product management" in t_lower or "product mgr" in t_lower:
            return True
        t_words = re.findall(r'\b\w+\b', t_lower)
        if "pm" in t_words or "pmt" in t_words:
            return True
        return False
        
    if company_name_lower == "google":
        for lvl in selected_levels_lower:
            if lvl == "product manager i":
                exclude_markers = ["ii", "senior", "sr.", "sr ", "group", "gpm", "director", "vp", "vice president", "lead", "principal", "staff"]
                if "pm i" in title_lower or "product manager i" in title_lower or (is_generic_pm_title(title_lower) and not any(m in title_lower for m in exclude_markers)):
                    return True
            elif lvl == "product manager ii":
                if "product manager ii" in title_lower or "pm ii" in title_lower or "l6" in title_lower:
                    return True
            elif lvl == "senior":
                if "senior" in title_lower or "sr." in title_lower or "sr " in title_lower:
                    return True
            elif lvl == "group":
                if "group" in title_lower or "gpm" in title_lower or "group product" in title_lower:
                    return True
            elif lvl == "director":
                if "director" in title_lower and "senior director" not in title_lower:
                    return True
            elif lvl == "senior director":
                if "senior director" in title_lower:
                    return True
            elif lvl == "vice president":
                if "vice president" in title_lower or "vp" in title_lower:
                    return True
                    
    elif company_name_lower == "meta":
        for lvl in selected_levels_lower:
            if lvl == "standard":
                exclude_markers = ["director", "vp", "vice president", "head of", "lead", "manager,", "group", "leadership"]
                if is_generic_pm_title(title_lower) and not any(m in title_lower for m in exclude_markers):
                    return True
            elif lvl == "leadership":
                include_markers = ["director", "vp", "vice president", "head of", "lead", "manager, product management", "group", "leadership"]
                if any(m in title_lower for m in include_markers):
                    return True
                    
    else:
        for lvl in selected_levels_lower:
            if lvl == "standard" or lvl.startswith("standard"):
                exclude_markers = ["senior", "sr.", "sr ", "principal", "staff", "director", "vp", "vice president", "lead", "associate", " ii", " iii", "head of", "chief", "partner", "group", "gpm"]
                if is_generic_pm_title(title_lower) and not any(m in title_lower for m in exclude_markers):
                    return True
            elif lvl == "senior":
                if "senior" in title_lower or "sr." in title_lower or "sr " in title_lower:
                    return True
            elif lvl == "staff":
                if "staff" in title_lower:
                    return True
            elif lvl == "group":
                if "group" in title_lower or "gpm" in title_lower or "group product" in title_lower:
                    return True
            elif lvl == "principal":
                if "principal" in title_lower:
                    return True
            elif lvl == "director":
                if "director" in title_lower and "senior director" not in title_lower:
                    return True
            elif lvl == "senior director":
                if "senior director" in title_lower:
                    return True
            elif lvl == "vice president":
                if "vice president" in title_lower or "vp" in title_lower:
                    return True
            elif lvl == "chief product officer" or lvl.startswith("chief product"):
                if "chief product" in title_lower or "cpo" in title_lower:
                    return True
                    
    return False

def matches_criteria(title, locations_list, search_criteria, company_name, config=None):
    title_lower = title.lower()
    
    # 1. PM Keyword Match (with robust abbreviation checking)
    keywords = search_criteria.get("custom_keywords", ["product manager", "product management", "technical product manager"])
    
    expanded_kws = []
    for kw in keywords:
        kw_l = kw.lower().strip()
        expanded_kws.append(kw_l)
        if kw_l == "product manager":
            expanded_kws.extend(["product mgr", "pm", "pmt"])
        elif kw_l == "technical product manager":
            expanded_kws.extend(["technical product mgr", "tpm", "pmt"])
            
    is_pm = False
    for kw in expanded_kws:
        if kw in ["pm", "tpm", "pmt"]:
            if re.search(r'\b' + re.escape(kw) + r'\b', title_lower):
                is_pm = True
                break
        else:
            if kw in title_lower:
                is_pm = True
                break
                
    if not is_pm:
        return False
        
    # 2. Level Match (Always check the company-specific level checkboxes!)
    if not title_matches_company_levels(title, company_name, config):
        return False
                
    # 3. Location Match
    locs = search_criteria.get("locations", ["New York", "NYC", "Remote"])
    is_loc = False
    for l in locations_list:
        if not l or l == "N/A":
            continue
        l_lower = l.lower().strip()
        for crit_loc in locs:
            crit_loc_lower = crit_loc.lower().strip()
            if crit_loc_lower == "remote":
                if "remote" in l_lower or "work from home" in l_lower or "anywhere" in l_lower:
                    is_loc = True
                    break
            elif crit_loc_lower in l_lower or l_lower in crit_loc_lower:
                is_loc = True
                break
        if is_loc:
            break
    return is_loc

# Crawler functions per platform
def check_greenhouse_jobs(board, company_name, cohort, search_criteria, config=None):
    url = f"https://api.greenhouse.io/v1/boards/{board}/jobs?content=true"
    leads = []
    try:
        r = requests.get(url, timeout=12)
        if r.status_code == 200:
            data = r.json()
            jobs = data.get("jobs", [])
            for job in jobs:
                title = job.get("title", "")
                loc = job.get("location", {}).get("name", "") if job.get("location") else ""
                
                if matches_criteria(title, [loc], search_criteria, company_name, config):
                    salary = "N/A"
                    content = job.get("content")
                    if content:
                        salary = extract_salary(content) or "N/A"
                    leads.append({
                        "cohort": cohort,
                        "company": company_name,
                        "role": title,
                        "location": loc,
                        "key_focus": "Discovered via Greenhouse automated crawl",
                        "url": job.get("absolute_url"),
                        "salary": salary
                    })
    except Exception as e:
        print(f"Error checking Greenhouse {board} ({company_name}): {e}")
    return leads

def check_ashby_jobs(board, company_name, cohort, search_criteria, config=None):
    url = f"https://api.ashbyhq.com/posting-api/job-board/{board}"
    leads = []
    try:
        r = requests.get(url, timeout=12)
        if r.status_code == 200:
            jobs = r.json().get("jobs", [])
            for j in jobs:
                title = j.get("title", "")
                loc = j.get("location", "")
                
                if matches_criteria(title, [loc], search_criteria, company_name, config):
                    salary = "N/A"
                    desc = j.get("descriptionPlain") or j.get("descriptionHtml")
                    if desc:
                        salary = extract_salary(desc) or "N/A"
                    leads.append({
                        "cohort": cohort,
                        "company": company_name,
                        "role": title,
                        "location": loc,
                        "key_focus": "Discovered via Ashby automated crawl",
                        "url": j.get("jobUrl"),
                        "salary": salary
                    })
    except Exception as e:
        print(f"Error checking Ashby {board} ({company_name}): {e}")
    return leads

def check_workable_jobs(board, company_name, cohort, search_criteria, config=None):
    url = f"https://apply.workable.com/api/v1/accounts/{board}/jobs"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Content-Type": "application/json"
    }
    leads = []
    try:
        r = requests.post(url, json={}, headers=headers, timeout=12)
        if r.status_code == 200:
            jobs = r.json().get("results", [])
            for j in jobs:
                title = j.get("title", "")
                loc_dict = j.get("location", {})
                city = loc_dict.get("city", "") if loc_dict else ""
                region = loc_dict.get("region", "") if loc_dict else ""
                loc_str = f"{city}, {region}" if city else "Remote"
                
                if matches_criteria(title, [loc_str], search_criteria, company_name, config):
                    leads.append({
                        "cohort": cohort,
                        "company": company_name,
                        "role": title,
                        "location": loc_str,
                        "key_focus": "Discovered via Workable automated crawl",
                        "url": f"https://apply.workable.com/{board}/j/{j.get('shortcode')}"
                    })
    except Exception as e:
        print(f"Error checking Workable {board} ({company_name}): {e}")
    return leads

def check_google_jobs(company_name, cohort, search_criteria, config=None):
    import urllib.parse
    import time
    leads = []
    
    if config is None:
        config = load_config()
        
    google_co = next((c for c in config.get("companies", []) if c.get("name", "").lower() == "google"), None)
    google_levels = google_co.get("levels") if google_co else None
    if google_levels is None:
        google_levels = ["Product Manager I", "Product Manager II", "Senior", "Group", "Director", "Senior Director", "Vice President"]
        
    keywords = search_criteria.get("custom_keywords", ["Product Manager"])
    
    # Build unique search queries
    queries = []
    for kw in keywords:
        for lvl in google_levels:
            lvl_lower = lvl.lower().strip()
            if lvl_lower in ["product manager i", "product manager ii"]:
                queries.append(kw)
            elif lvl_lower == "senior":
                queries.append(f"Senior {kw}")
            elif lvl_lower == "group":
                queries.append(f"Group {kw}")
            elif lvl_lower == "director":
                queries.append(f"Director {kw}")
            elif lvl_lower == "senior director":
                queries.append(f"Senior Director {kw}")
            elif lvl_lower == "vice president":
                queries.append(f"Vice President {kw}")
                
    # Deduplicate queries while preserving order
    seen_queries = set()
    unique_queries = []
    for q in queries:
        q_norm = q.lower().strip()
        if q_norm not in seen_queries:
            seen_queries.add(q_norm)
            unique_queries.append(q)
            
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    seen_job_ids = set()
    
    for query_term in unique_queries:
        # We query for New York, NY, USA as location to avoid global result dilution
        url = f"https://www.google.com/about/careers/applications/jobs/results?q={urllib.parse.quote(query_term)}&location=New%20York%2C%20NY%2C%20USA"
        try:
            r = requests.get(url, headers=headers, timeout=12)
            if r.status_code != 200:
                print(f"  Google Careers: failed to fetch query '{query_term}', status {r.status_code}")
                continue
                
            text_content = r.text  # Cache string to prevent loop decoding overhead
            idx = text_content.find("key: 'ds:1'")
            if idx == -1:
                idx = text_content.find('key: "ds:1"')
                
            if idx != -1:
                data_idx = text_content.find("data:", idx)
                if data_idx != -1:
                     start_array = text_content.find("[", data_idx)
                     if start_array != -1:
                        bracket_count = 0
                        end_array = -1
                        total_len = len(text_content)
                        
                        for i in range(start_array, total_len):
                            char = text_content[i]
                            if char == '[':
                                bracket_count += 1
                            elif char == ']':
                                bracket_count -= 1
                                if bracket_count == 0:
                                    end_array = i + 1
                                    break
                                    
                        if end_array != -1:
                            json_str = text_content[start_array:end_array]
                            data = json.loads(json_str)
                            jobs_list = data[0] if data and len(data) > 0 else []
                            
                            for job in jobs_list:
                                job_id = job[0]
                                if job_id in seen_job_ids:
                                    continue
                                    
                                title = job[1]
                                locs = [loc[0] for loc in job[9]] if len(job) > 9 and job[9] else []
                                
                                if matches_criteria(title, locs, search_criteria, company_name, config):
                                    seen_job_ids.add(job_id)
                                    job_url = f"https://www.google.com/about/careers/applications/jobs/results/{job_id}"
                                    leads.append({
                                        "cohort": cohort,
                                        "company": company_name,
                                        "role": title,
                                        "location": ", ".join(locs),
                                        "key_focus": f"Google Careers HTML crawl ({query_term})",
                                        "url": job_url,
                                        "salary": "N/A"
                                    })
            time.sleep(0.5)
        except Exception as e:
            print(f"  Error checking Google Careers for '{query_term}': {e}")
            
    return leads


def check_microsoft_jobs(company_name, cohort, search_criteria, config=None):
    """
    Query Microsoft's Eightfold PCSX career API to find PM roles.
    Steps: 1) Get CSRF token from landing page, 2) Paginate GET /api/pcsx/search,
    3) Fetch position_details for salary, 4) Return leads with individual job URLs.
    """
    import time as _time
    leads = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    }
    session = requests.Session()
    
    try:
        # Step 1: Get CSRF token from landing page
        landing_url = "https://apply.careers.microsoft.com/careers?query=product+manager&start=0&location=New+York%2C++NY%2C++United+States&sort_by=relevance&filter_distance=160&filter_include_remote=1"
        resp = session.get(landing_url, headers=headers, timeout=15)
        csrf_match = re.search(r'name="_csrf"\s+content="([^"]+)"', resp.text)
        csrf_token = csrf_match.group(1) if csrf_match else ""
        
        api_headers = {
            "User-Agent": headers["User-Agent"],
            "Accept": "application/json",
            "X-CSRF-Token": csrf_token,
            "Referer": landing_url,
        }
        
        # Step 2: Paginate through search results
        search_url = "https://apply.careers.microsoft.com/api/pcsx/search"
        all_positions = []
        start = 0
        page_size = 10
        
        while True:
            params = {
                "domain": "microsoft.com",
                "query": "product manager",
                "location": "New York, NY, United States",
                "start": start,
                "num": page_size,
                "sort_by": "relevance",
                "filter_distance": 160,
                "filter_include_remote": 1,
            }
            
            resp2 = session.get(search_url, params=params, headers=api_headers, timeout=15)
            if resp2.status_code != 200:
                print(f"  Microsoft API error at start={start}: {resp2.status_code}")
                break
            
            data = resp2.json()
            positions = data.get('data', {}).get('positions', [])
            total = data.get('data', {}).get('count', 0)
            
            if not positions:
                break
            
            all_positions.extend(positions)
            start += page_size
            if start >= total:
                break
            _time.sleep(0.3)
        
        print(f"  Microsoft: fetched {len(all_positions)}/{total if 'total' in dir() else '?'} total search results")
        
        # Step 3: Filter for PM roles and build leads with salary
        for pos in all_positions:
            title = pos.get('name', '')
            pid = pos.get('id', '')
            locations = pos.get('locations', [])
            loc_str = '; '.join(locations) if locations else 'N/A'
            
            # Filter using criteria matcher
            if not matches_criteria(title, locations, search_criteria, company_name, config):
                continue
            
            job_url = f"https://apply.careers.microsoft.com/careers/job/{pid}"
            
            # Fetch salary from position details
            salary = 'N/A'
            try:
                detail_url = f"https://apply.careers.microsoft.com/api/pcsx/position_details?domain=microsoft.com&position_id={pid}"
                resp3 = session.get(detail_url, headers=api_headers, timeout=10)
                if resp3.status_code == 200:
                    desc = resp3.json().get('data', {}).get('jobDescription', '')
                    salary_match = re.search(r'\$[\d,]+\s*[-\u2013to]+\s*\$[\d,]+(?:\s*(?:per\s+)?(?:year|annually|USD))?', desc)
                    if salary_match:
                        salary = salary_match.group(0)
                    elif not salary:
                        salary = extract_salary(desc)
            except Exception:
                pass
            _time.sleep(0.2)
            
            leads.append({
                "cohort": cohort,
                "company": company_name,
                "role": title,
                "location": loc_str,
                "key_focus": "Microsoft PCSX API",
                "url": job_url,
                "salary": salary,
            })
    except Exception as e:
        print(f"Error checking Microsoft: {e}")
    return leads

def check_apple_jobs(company_name, cohort, search_criteria, config=None):
    leads = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    # Get first query term from settings, default to "Product Manager"
    keywords = search_criteria.get("custom_keywords", ["Product Manager"])
    query_term = keywords[0] if keywords else "Product Manager"
    query_quoted = urllib.parse.quote(query_term)
    
    # Fetch NY State search page 1 and general search pages 1-2 to capture all relevant roles
    urls_to_crawl = [
        f"https://jobs.apple.com/en-us/search?search={query_quoted}&sort=relevance&location=new-york-state985&page=1",
        f"https://jobs.apple.com/en-us/search?search={query_quoted}&sort=relevance&page=1",
        f"https://jobs.apple.com/en-us/search?search={query_quoted}&sort=relevance&page=2"
    ]
    
    seen_urls = set()
    
    for url in urls_to_crawl:
        try:
            r = requests.get(url, headers=headers, timeout=12)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                items = soup.find_all('li', class_='rc-accordion-item')
                for li in items:
                    a_tag = li.find('a', class_='link-inline')
                    if a_tag:
                        title = a_tag.get_text(strip=True)
                        href = a_tag['href']
                        job_url = f"https://jobs.apple.com{href}"
                        
                        if job_url not in seen_urls:
                            # Extract location
                            loc_div = li.find(class_='job-title-location')
                            location = "N/A"
                            if loc_div:
                                spans = loc_div.find_all('span')
                                location = " ".join(s.get_text(strip=True) for s in spans if 'a11y' not in s.get('class', []))
                            
                            # Match criteria
                            if matches_criteria(title, [location], search_criteria, company_name, config):
                                seen_urls.add(job_url)
                                
                                # Fetch salary from React hydration data
                                salary = "N/A"
                                try:
                                    r_det = requests.get(job_url, headers=headers, timeout=10)
                                    if r_det.status_code == 200:
                                        match_hyd = re.search(r'window\.__staticRouterHydrationData\s*=\s*JSON\.parse\("((?:[^"\\]|\\.)*)"\);', r_det.text)
                                        if match_hyd:
                                            escaped_json = match_hyd.group(1)
                                            raw_json_str = json.loads(f'"{escaped_json}"')
                                            hyd_data = json.loads(raw_json_str)
                                            jobs_data = hyd_data.get("loaderData", {}).get("jobDetails", {}).get("jobsData", {})
                                            
                                            text_chunks = []
                                            text_chunks.append(jobs_data.get("description", ""))
                                            text_chunks.append(jobs_data.get("jobSummary", ""))
                                            for foot in jobs_data.get("postingFooters", []):
                                                for loc_key, loc_list in foot.get("localizations", {}).items():
                                                    for loc_item in loc_list:
                                                        text_chunks.append(loc_item.get("content", ""))
                                            
                                            full_desc = " ".join(text_chunks)
                                            salary = extract_salary(full_desc) or "N/A"
                                except Exception as e:
                                    print(f"Error fetching Apple salary for {title}: {e}")
                                    
                                leads.append({
                                    "cohort": cohort,
                                    "company": company_name,
                                    "role": title,
                                    "location": location,
                                    "key_focus": "Apple Careers HTML crawl",
                                    "url": job_url,
                                    "salary": salary
                                })
        except Exception as e:
            print(f"Error crawling Apple URL {url}: {e}")
            
    return leads

def check_amazon_jobs(company_name, cohort, search_criteria, config=None):
    leads = []
    url = "https://www.amazon.jobs/en/search.json"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    if config is None:
        config = load_config()
        
    amazon_co = next((c for c in config.get("companies", []) if c.get("name", "").lower() == "amazon"), None)
    amazon_levels = amazon_co.get("levels") if amazon_co else None
    if amazon_levels is None:
        amazon_levels = ["Standard", "Senior", "Staff", "Principal", "Director", "Senior Director", "Vice President", "Chief Product Officer"]
        
    queries = []
    for lvl in amazon_levels:
        lvl_lower = lvl.lower().strip()
        if lvl_lower == "standard":
            queries.append("product manager")
        elif lvl_lower == "senior":
            queries.append("senior product manager")
        elif lvl_lower == "principal":
            queries.append("principal product manager")
        elif lvl_lower == "staff":
            queries.append("staff product manager")
        elif lvl_lower == "director":
            queries.append("director product manager")
        elif lvl_lower == "vice president":
            queries.append("vp product manager")
            
    seen_queries = set()
    unique_queries = []
    for q in queries:
        if q not in seen_queries:
            seen_queries.add(q)
            unique_queries.append(q)
            
    if not unique_queries:
        unique_queries = ["product manager"]
        
    seen_job_ids = set()
    
    for base_query in unique_queries:
        params = {
            "offset": 0,
            "result_limit": 100,
            "sort": "relevant",
            "loc_group_id": "new-york-metro-area",
            "base_query": base_query
        }
        try:
            r = requests.get(url, params=params, headers=headers, timeout=12)
            if r.status_code == 200:
                jobs = r.json().get("jobs", [])
                for j in jobs:
                    job_id = j.get("id")
                    if job_id in seen_job_ids:
                        continue
                    seen_job_ids.add(job_id)
                    
                    title = j.get("title", "")
                    loc = j.get("normalized_location", "")
                    
                    # Verify location and level/keywords matching
                    if matches_criteria(title, [loc], search_criteria, company_name, config):
                        job_url = f"https://www.amazon.jobs{j.get('job_path')}"
                        leads.append({
                            "cohort": cohort,
                            "company": company_name,
                            "role": title,
                            "location": loc,
                            "key_focus": f"Amazon Public Search API ({base_query})",
                            "url": job_url
                        })
        except Exception as e:
            print(f"Error checking Amazon query '{base_query}': {e}")
            
    return leads

def check_meta_jobs(company_name, cohort, search_criteria, config=None):
    import time
    leads = []
    session = requests.Session()
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Sec-Ch-Ua": '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1"
    }
    
    try:
        # Step 1: GET jobsearch to fetch cookies and extract LSD token
        landing_url = "https://www.metacareers.com/jobsearch?q=product%20manager&offices[0]=New%20York%2C%20NY"
        r = session.get(landing_url, headers=headers, timeout=12)
        if r.status_code != 200:
            print(f"Meta Careers GET request failed: {r.status_code}")
            return leads
            
        lsd_match = re.search(r'"LSD"\s*,\s*\[\],\s*\{"token"\s*:\s*["\']([^"\']+)["\']', r.text)
        if not lsd_match:
            lsd_match = re.search(r'"lsd"\s*:\s*["\']([^"\']+)["\']', r.text)
        
        if not lsd_match:
            print("Meta Scraper: LSD token not found.")
            return leads
            
        lsd_token = lsd_match.group(1)
        
        # Step 2: POST to GraphQL
        api_url = "https://www.metacareers.com/api/graphql/"
        api_headers = {
            "User-Agent": headers["User-Agent"],
            "Accept": "*/*",
            "Accept-Language": "en-US,en;q=0.9",
            "Content-Type": "application/x-www-form-urlencoded",
            "Sec-Ch-Ua": headers["Sec-Ch-Ua"],
            "Sec-Ch-Ua-Mobile": headers["Sec-Ch-Ua-Mobile"],
            "Sec-Ch-Ua-Platform": headers["Sec-Ch-Ua-Platform"],
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "Origin": "https://www.metacareers.com",
            "Referer": landing_url,
            "X-FB-LSD": lsd_token
        }
        
        payload = {
            "lsd": lsd_token,
            "doc_id": "26703205452636175",
            "variables": json.dumps({
                "search_input": {
                    "q": "product manager",
                    "offices": ["New York, NY"],
                    "divisions": [],
                    "roles": [],
                    "leadership_levels": [],
                    "saved_jobs": [],
                    "saved_searches": [],
                    "sub_teams": [],
                    "teams": [],
                    "is_leadership": False,
                    "is_remote_only": False,
                    "sort_by_new": False,
                    "page": 1
                }
            })
        }
        
        r2 = session.post(api_url, data=payload, headers=api_headers, timeout=12)
        if r2.status_code == 200:
            data = r2.json()
            search_res = data.get("data", {}).get("job_search_with_featured_jobs_v2", {})
            all_jobs = search_res.get("all_jobs", [])
            featured_jobs = search_res.get("featured_jobs", [])
            total_jobs = all_jobs + featured_jobs
            
            seen_ids = set()
            for j in total_jobs:
                jid = j.get("id")
                if not jid or jid in seen_ids:
                    continue
                seen_ids.add(jid)
                
                title = j.get("title", "")
                locations = j.get("locations", [])
                
                if matches_criteria(title, locations, search_criteria, company_name, config):
                    job_url = f"https://www.metacareers.com/jobs/{jid}/"
                    
                    # Fetch salary from details page
                    salary = "N/A"
                    try:
                        detail_headers = headers.copy()
                        detail_headers["Referer"] = landing_url
                        r3 = session.get(job_url, headers=detail_headers, timeout=8)
                        if r3.status_code == 200:
                            min_match = re.search(r'"compensation_amount_minimum"\s*:\s*["\']([^"\']+)["\']', r3.text)
                            max_match = re.search(r'"compensation_amount_maximum"\s*:\s*["\']([^"\']+)["\']', r3.text)
                            if min_match and max_match:
                                min_val = min_match.group(1).replace('\\/', '/')
                                max_val = max_match.group(1).replace('\\/', '/')
                                salary = f"{min_val} - {max_val}"
                    except Exception as ex:
                        print(f"Error fetching Meta salary details for job {jid}: {ex}")
                    time.sleep(0.2)
                    
                    leads.append({
                        "cohort": cohort,
                        "company": company_name,
                        "role": title,
                        "location": ", ".join(locations),
                        "key_focus": "Meta GraphQL API",
                        "url": job_url,
                        "salary": salary
                    })
    except Exception as e:
        print(f"Error checking Meta: {e}")
    return leads

def check_nvidia_jobs(board, company_name, cohort, search_criteria, config=None):
    leads = []
    
    base_url = "https://jobs.nvidia.com/careers"
    api_url = "https://jobs.nvidia.com/api/pcsx/search"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    try:
        session = requests.Session()
        r = session.get(base_url, headers=headers, timeout=12)
        if r.status_code != 200:
            return leads
            
        soup = BeautifulSoup(r.text, 'html.parser')
        csrf_tag = soup.find('meta', attrs={'name': '_csrf'})
        csrf_token = csrf_tag.get('content') if csrf_tag else None
        
        if not csrf_token:
            csrf_token = r.headers.get("X-CSRF-Token")
            
        if not csrf_token:
            print("Nvidia Scraper: CSRF token not found.")
            return leads
            
        get_headers = {
            "User-Agent": headers["User-Agent"],
            "Accept": "application/json",
            "X-CSRF-Token": csrf_token,
            "Referer": base_url
        }
        
        keywords = search_criteria.get("custom_keywords", ["Product Manager"])
        query_term = keywords[0] if keywords else "Product Manager"
        
        locs = search_criteria.get("locations", ["New York", "NYC", "Remote"])
        seen_pos_ids = set()
        
        for loc in locs:
            params = {
                "domain": "nvidia.com",
                "query": query_term,
                "location": loc,
                "limit": 100,
                "start": 0
            }
            
            r_api = session.get(api_url, params=params, headers=get_headers, timeout=12)
            if r_api.status_code == 200:
                data = r_api.json()
                inner_data = data.get("data", {})
                positions = inner_data.get("positions", [])
                for pos in positions:
                    title = pos.get("name")
                    pos_id = pos.get("id")
                    locations = pos.get("locations", [])
                    
                    if title and pos_id and pos_id not in seen_pos_ids:
                        if matches_criteria(title, locations, search_criteria, company_name, config):
                            seen_pos_ids.add(pos_id)
                            job_url = f"https://jobs.nvidia.com/careers?pid={pos_id}"
                            
                            # Fetch salary from Nvidia Eightfold API details
                            salary = "N/A"
                            try:
                                detail_url = f"https://jobs.nvidia.com/api/pcsx/position_details?domain=nvidia.com&position_id={pos_id}"
                                r_det = session.get(detail_url, headers=get_headers, timeout=10)
                                if r_det.status_code == 200:
                                    desc = r_det.json().get("data", {}).get("jobDescription", "")
                                    salary = extract_salary(desc) or "N/A"
                            except Exception as e:
                                print(f"Error fetching Nvidia salary for {title}: {e}")
                                
                            leads.append({
                                "cohort": cohort,
                                "company": company_name,
                                "role": title,
                                "location": ", ".join(locations),
                                "key_focus": "Nvidia Eightfold API crawl",
                                "url": job_url,
                                "salary": salary
                            })
            else:
                print(f"Nvidia Scraper API request failed for location '{loc}': {r_api.status_code}")
    except Exception as e:
        print(f"Error checking Nvidia: {e}")
        
    return leads

def check_netflix_jobs(company_name, cohort, search_criteria, config=None):
    import html
    leads = []
    
    # Get first query term from settings, default to "Product Manager"
    keywords = search_criteria.get("custom_keywords", ["Product Manager"])
    query_term = keywords[0] if keywords else "Product Manager"
    
    # Query Netflix Careers (Phenom) pre-filtered for NYC/US
    loc_query = "New York, NY, United States"
    url = f"https://explore.jobs.netflix.net/careers?query={urllib.parse.quote(query_term)}&location={urllib.parse.quote(loc_query)}&domain=netflix.com&sort_by=relevance"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    try:
        r = requests.get(url, headers=headers, timeout=12)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            code_tag = soup.find('code', id='smartApplyData')
            if code_tag:
                decoded = html.unescape(code_tag.get_text())
                config_loaded = json.loads(decoded)
                positions = config_loaded.get("positions", [])
                for pos in positions:
                    title = pos.get("name") or pos.get("jobTitle", "")
                    pos_id = pos.get("id") or pos.get("postingId")
                    locs_list = pos.get("locations", [])
                    
                    if pos_id and title:
                        if matches_criteria(title, locs_list, search_criteria, company_name, config):
                            job_url = f"https://explore.jobs.netflix.net/careers?pid={pos_id}"
                            
                            # Fetch salary from JSON-LD
                            salary = "N/A"
                            try:
                                r_det = requests.get(job_url, headers=headers, timeout=10)
                                if r_det.status_code == 200:
                                    det_soup = BeautifulSoup(r_det.text, 'html.parser')
                                    ld_scripts = det_soup.find_all('script', type='application/ld+json')
                                    for s in ld_scripts:
                                        if s.string:
                                            try:
                                                ld_data = json.loads(s.string)
                                                if ld_data.get("@type") == "JobPosting":
                                                    desc = ld_data.get("description", "")
                                                    salary = extract_salary(desc) or "N/A"
                                                    break
                                            except:
                                                pass
                            except Exception as e:
                                print(f"Error fetching Netflix salary for {title}: {e}")
                                
                            leads.append({
                                "cohort": cohort,
                                "company": company_name,
                                "role": title,
                                "location": ", ".join(locs_list),
                                "key_focus": "Netflix Phenom Portal crawl",
                                "url": job_url,
                                "salary": salary
                            })
    except Exception as e:
        print(f"Error checking Netflix: {e}")
    return leads

# Fallback Generic DuckDuckGo Search Crawler
def check_company_jobs_ddg(company_name, cohort, search_criteria, config=None):
    leads = []
    locs = search_criteria.get("locations", ["New York", "NYC", "Remote"])
    loc_clause = " OR ".join(f'"{loc}"' for loc in locs)
    
    portal_host = None
    if config:
        for co in config.get("companies", []):
            if co.get("name", "").lower() == company_name.lower():
                portal_url = co.get("portal_url", "")
                if portal_url:
                    try:
                        parsed = urllib.parse.urlparse(portal_url)
                        if parsed.netloc:
                            portal_host = parsed.netloc.lower()
                            if portal_host.startswith("www."):
                                portal_host = portal_host[4:]
                    except Exception:
                        pass
                break

    # Customize queries based on company
    if company_name.lower() == "meta":
        query = f'site:metacareers.com/jobs "Product Manager" ({loc_clause})'
    elif company_name.lower() == "tesla":
        query = f'site:tesla.com/careers "Product Manager" ({loc_clause})'
    elif portal_host and not any(x in portal_host for x in ["google.com", "microsoft.com", "amazon.jobs", "amazon.com", "netflix.com", "nvidia.com", "apple.com", "greenhouse.io", "ashbyhq.com", "workable.com"]):
        # Search the custom careers portal host AND standard boards
        query = f'"{company_name}" "Product Manager" ({loc_clause}) (site:{portal_host} OR site:lever.co OR site:greenhouse.io OR site:ashbyhq.com)'
    else:
        query = f'"{company_name}" "Product Manager" ({loc_clause}) site:lever.co OR site:greenhouse.io OR site:ashbyhq.com'
        
    try:
        urls = search_ddg(query)
        for url in urls:
            if not url or is_placeholder_url(url):
                continue
                
            # Filter URLs
            url_lower = url.lower()
            is_job_page = False
            if company_name.lower() == "meta" and ("/jobs/" in url_lower or "/v2/jobs/" in url_lower):
                is_job_page = True
            elif company_name.lower() == "tesla" and "/job/" in url_lower:
                is_job_page = True
            elif any(x in url_lower for x in ["/jobs/", "/job/", "/posting/", "/careers/", "/position/", "/positions/"]):
                is_job_page = True
            elif portal_host and portal_host in url_lower:
                if any(x in url_lower for x in ["search", "results", "category", "location", "jobs-results"]):
                    is_job_page = False
                else:
                    is_job_page = True
                
            if is_job_page:
                # Fetch page to read the title
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                }
                r = requests.get(url, headers=headers, timeout=6)
                if r.status_code == 200:
                    soup = BeautifulSoup(r.text, 'html.parser')
                    title_text = soup.title.string if soup.title else ""
                    if title_text:
                        title_clean = title_text.split(" - ")[0].split(" | ")[0].split(" at ")[0].strip()
                        
                        # Verify title matches criteria
                        if matches_criteria(title_clean, [title_text], search_criteria, company_name, config):
                            # Map location
                            loc_found = "New York, NY" if ("new york" in title_text.lower() or "nyc" in title_text.lower()) else "Remote / Various"
                            leads.append({
                                "cohort": cohort,
                                "company": company_name,
                                "role": title_clean,
                                "location": loc_found,
                                "key_focus": "DDG Indexed Crawler search",
                                "url": url
                            })
    except Exception as e:
        print(f"Error search-crawling {company_name}: {e}")
    return leads

def search_ddg_snippets(query):
    url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(query)}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    for attempt in range(3):
        try:
            r = requests.get(url, headers=headers, timeout=10)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                snippets = []
                for a in soup.find_all('a', class_='result__snippet'):
                    snippets.append(a.get_text())
                if snippets:
                    return snippets
            elif r.status_code == 202:
                soup = BeautifulSoup(r.text, 'html.parser')
                meta_refresh = soup.find('meta', attrs={'http-equiv': 'refresh'})
                if meta_refresh:
                    content = meta_refresh.get('content', '')
                    if 'url=' in content:
                        refresh_url = content.split('url=')[-1]
                        if refresh_url.startswith('/'):
                            refresh_url = f"https://html.duckduckgo.com{refresh_url}"
                        url = refresh_url
                        continue
        except Exception as e:
            print(f"Error querying DDG for '{query}': {e}")
        import time
        time.sleep(1)
    return []

def fetch_company_details(name):
    print(f"Enriching details for: {name}")
    details = {
        "hq": "N/A",
        "founded": "N/A",
        "revenue": "N/A",
        "employees": "N/A",
        "domain": "N/A"
    }
    
    # Query: general profile
    snippets = search_ddg_snippets(f"{name} company profile headquarters founded year employees revenue")
    full_text = " ".join(snippets)
    
    # 1. Extract Founded Year
    founded_match = re.search(r'\bfounded\s+(?:in\s+)?([12][0-9]{3})\b', full_text, re.IGNORECASE)
    if not founded_match:
        founded_match = re.search(r'\bestablished\s+(?:in\s+)?([12][0-9]{3})\b', full_text, re.IGNORECASE)
    if not founded_match:
        years = re.findall(r'\b(19[789][0-9]|20[0-2][0-9])\b', full_text)
        if years:
            details["founded"] = years[0]
    else:
        details["founded"] = founded_match.group(1)
        
    # 2. Extract HQ Location
    hq_match = re.search(r'\bheadquartered\s+in\s+([^,\.\n]{2,30},?\s*[A-Z]{2})', full_text, re.IGNORECASE)
    if not hq_match:
        hq_match = re.search(r'\bheadquarters\s+(?:is\s+in|in|are\s+in)?\s*([^,\.\n]{2,30},?\s*[A-Z]{2})', full_text, re.IGNORECASE)
    if hq_match:
        details["hq"] = hq_match.group(1).strip()
    else:
        # Try a different query
        snippets_hq = search_ddg_snippets(f"{name} company headquarters location")
        text_hq = " ".join(snippets_hq)
        hq_match2 = re.search(r'\b(?:headquartered|headquarters)\s+(?:in|is\s+in|are\s+in|at)?\s*([^,\.\n]{2,30},?\s*[A-Za-z]{2,20})', text_hq, re.IGNORECASE)
        if hq_match2:
            details["hq"] = hq_match2.group(1).strip()
            
    # 3. Extract Employees
    emp_match = re.search(r'\b([0-9,]+(?:\s*\+)?)\s*employees\b', full_text, re.IGNORECASE)
    if not emp_match:
        emp_match = re.search(r'\bhas\s+(?:over|around|about|approximately)?\s*([0-9,k\+]+)\s*employees\b', full_text, re.IGNORECASE)
    if emp_match:
        details["employees"] = emp_match.group(1).strip()
    else:
        snippets_emp = search_ddg_snippets(f"{name} number of employees count")
        text_emp = " ".join(snippets_emp)
        emp_match2 = re.search(r'\b([0-9,k\+]{2,10})\s*employees\b', text_emp, re.IGNORECASE)
        if emp_match2:
            details["employees"] = emp_match2.group(1).strip()

    # 4. Extract Revenue
    rev_match = re.search(r'\b(\$[0-9\.]+(?:\s*[BMbm]illion|[BMbm]?)?)\s*(?:in\s+)?revenue\b', full_text, re.IGNORECASE)
    if not rev_match:
        rev_match = re.search(r'\bannual\s+revenue\s*(?:of|is)?\s*(\$[0-9\.]+(?:\s*[BMbm]illion|[BMbm]?)?)', full_text, re.IGNORECASE)
    if rev_match:
        details["revenue"] = rev_match.group(1).strip()
    else:
        snippets_rev = search_ddg_snippets(f"{name} annual revenue")
        text_rev = " ".join(snippets_rev)
        rev_match2 = re.search(r'(\$[0-9\.]+(?:\s*[BMbm]illion|[BMbm]?)?)', text_rev)
        if rev_match2:
            details["revenue"] = rev_match2.group(1).strip()

    # 5. Extract Domain
    if snippets:
        first_snip = snippets[0].lower()
        domain_match = re.search(rf'{name.lower()}\s+(?:is\s+a|is\s+an)\s+([^,\.]+)', first_snip)
        if domain_match:
            details["domain"] = domain_match.group(1).strip().capitalize()
        else:
            snippets_ind = search_ddg_snippets(f"{name} company industry domain")
            if snippets_ind:
                details["domain"] = snippets_ind[0][:80].strip() + "..."
                
    return details

def run_daily_scan(target_company=None):
    if target_company:
        print(f"Initiating single company job leads scan for '{target_company}'...")
    else:
        print("Initiating daily job leads scan...")
    config = load_config()
    search_criteria = config.get("search_criteria", {})
    companies = config.get("companies", [])
    
    # Auto-enrich company details if needed
    config_changed = False
    for co in companies:
        if target_company and co["name"].lower() != target_company.lower():
            continue
            
        if (co.get("hq") == "N/A" or co.get("founded") == "N/A" or 
            co.get("revenue") == "N/A" or co.get("employees") == "N/A" or 
            co.get("domain") == "N/A" or not co.get("hq") or not co.get("domain")):
            
            print(f"Auto-enriching details for {co['name']}...")
            try:
                enriched = fetch_company_details(co["name"])
                # Update fields that are N/A or empty
                for field in ["hq", "founded", "revenue", "employees", "domain"]:
                    if co.get(field) == "N/A" or not co.get(field):
                        co[field] = enriched[field]
                config_changed = True
            except Exception as e:
                print(f"Error enriching company details for {co['name']}: {e}")
            
    if config_changed:
        print("Saving enriched company details back to config.json...")
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Error saving enriched config: {e}")
    
    if not os.path.exists(TRACKER_PATH):
        print("Tracker Excel file not found. Creating a blank tracker first.")
        # Create sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Leads"
        headers = ['Select to Apply', 'Cohort', 'Company', 'Role', 'Location', 'Key Focus / Technical Area', 'Status', 'Specific Job Link', 'Date Added', 'Notes', 'Compensation']
        ws.append(headers)
        wb.save(TRACKER_PATH)
        
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb["Job Leads"]
    
    # Ensure Column 15 has the header "Archive Reason"
    if ws.max_column < 15 or ws.cell(row=1, column=15).value is None:
        ws.cell(row=1, column=15).value = "Archive Reason"
        
    # Read existing leads to prevent duplicates
    existing_keys = {}
    for row in range(2, ws.max_row + 1):
        company = ws.cell(row=row, column=3).value
        role = ws.cell(row=row, column=4).value
        if company and role:
            key = (normalize_string(company), normalize_string(role))
            existing_keys[key] = row
            
    all_new_leads = []
    scanned_companies = set()
    crawled_keys = set()
    
    # Loop over all configured companies
    for co in companies:
        name = co["name"]
        if target_company and name.lower() != target_company.lower():
            continue
        platform = co["platform"]
        board_id = co["board_id"]
        cohort = co["cohort"]
        capability = co.get("capability", "active_sync")
        
        # Check if all levels are de-selected for this company
        if not co.get("levels"):
            print(f"Skipping {name} (no target levels selected)...")
            continue

        if capability == "portal_only":
            print(f"Skipping {name} (portal_only: manual check required)...")
            continue
            
        print(f"Scanning {name} using platform: {platform}...")
        co_leads = []
        
        try:
            if platform == "google":
                co_leads = check_google_jobs(name, cohort, search_criteria, config)
            elif platform == "microsoft":
                co_leads = check_microsoft_jobs(name, cohort, search_criteria, config)
            elif platform == "netflix":
                co_leads = check_netflix_jobs(name, cohort, search_criteria, config)
            elif platform == "apple":
                co_leads = check_apple_jobs(name, cohort, search_criteria, config)
            elif platform == "amazon":
                co_leads = check_amazon_jobs(name, cohort, search_criteria, config)
            elif platform == "meta":
                co_leads = check_meta_jobs(name, cohort, search_criteria, config)
            elif platform == "nvidia":
                co_leads = check_nvidia_jobs(board_id, name, cohort, search_criteria, config)
            elif platform == "greenhouse":
                co_leads = check_greenhouse_jobs(board_id, name, cohort, search_criteria, config)
            elif platform == "ashby":
                co_leads = check_ashby_jobs(board_id, name, cohort, search_criteria, config)
            elif platform == "workable":
                co_leads = check_workable_jobs(board_id, name, cohort, search_criteria, config)
            else: # tesla, ddg, or any other general
                co_leads = check_company_jobs_ddg(name, cohort, search_criteria, config)
                
            scanned_companies.add(normalize_string(name))
            print(f" - Found {len(co_leads)} matching roles for {name}.")
            all_new_leads.extend(co_leads)
            
            for lead in co_leads:
                crawled_keys.add((normalize_string(lead["company"]), normalize_string(lead["role"])))
        except Exception as e:
            print(f"Error during scan of {name}: {e}")
            
    # Styles for Excel
    data_font = Font(name="Segoe UI", size=10)
    link_font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    today_str = datetime.today().strftime('%Y-%m-%d')
    current_row = ws.max_row + 1
    
    updated_count = 0
    added_count = 0
    
    for lead in all_new_leads:
        key = (normalize_string(lead["company"]), normalize_string(lead["role"]))
        
        # Check duplicates
        if key in existing_keys:
            row_idx = existing_keys[key]
            cell = ws.cell(row=row_idx, column=8)
            existing_url = cell.hyperlink.target if cell.hyperlink else cell.value
            
            if is_placeholder_url(existing_url):
                print(f"Updating URL for: {lead['company']} - {lead['role']}")
                cell.value = "Link to Job Description"
                cell.hyperlink = lead["url"]
                cell.font = link_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row_idx, column=10).value = "Direct link auto-updated via scan."
                updated_count += 1
            
            # If the job was previously archived as Closed, but shows up again, restore it
            current_status = ws.cell(row=row_idx, column=7).value
            if current_status == "Archived" and ws.cell(row=row_idx, column=15).value == "Closed":
                print(f"Restoring reopened job: {lead['company']} - {lead['role']}")
                ws.cell(row=row_idx, column=7).value = "Lead"
                ws.cell(row=row_idx, column=15).value = ""
                ws.cell(row=row_idx, column=10).value = "Reopened and restored via daily scan."
                updated_count += 1
            continue
            
        # Extract salary for new lead (use pre-fetched if available)
        salary = lead.get("salary", "N/A") or "N/A"
        if salary == "N/A" and lead.get("url") and not is_placeholder_url(lead["url"]):
            try:
                headers_req = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                }
                res = requests.get(lead["url"], headers=headers_req, timeout=10)
                if res.status_code == 200:
                    salary = extract_salary(res.text) or "N/A"
            except Exception as e:
                print(f"Error fetching salary for {lead['company']} - {lead['role']}: {e}")
        
        # Write new lead row cell-by-cell (NOT ws.append, which causes hyperlink offset bugs)
        row_data = [
            "[ ]", 
            lead["cohort"], 
            lead["company"], 
            lead["role"], 
            lead["location"], 
            lead["key_focus"], 
            "Lead", 
            None,  # Column 8 placeholder - URL set via hyperlink below
            today_str, 
            "Discovered by daily scan.",
            salary
        ]
        for col_num, val in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col_num).value = val
        
        ws.row_dimensions[current_row].height = 22
        
        # Set hyperlink explicitly on column 8 BEFORE any other cell operations
        url_cell = ws.cell(row=current_row, column=8)
        url_cell.value = "Link to Job Description"
        url_cell.hyperlink = lead["url"]
        url_cell.font = link_font
        url_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            if col_num == 8:
                continue  # Already handled above
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            if col_num == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Segoe UI", size=10, bold=True)
                
            if col_num in [5, 7, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            if col_num == 2:
                # Color code cohorts
                cohort = cell.value
                if cohort == "Mag 7":
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Light Blue
                elif cohort == "AI Labs":
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green
                elif cohort == "High-Growth Startups":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light Yellow
                elif cohort == "Non-Mag 7":
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Light Peach
                
        print(f"Added: {lead['company']} - {lead['role']}")
        current_row += 1
        added_count += 1
        
    # Auto-archive closed roles or roles that no longer match the current criteria
    archived_closed_count = 0
    archived_criteria_count = 0
    for row_idx in range(2, ws.max_row + 1):
        row_company = ws.cell(row=row_idx, column=3).value
        row_role = ws.cell(row=row_idx, column=4).value
        row_status = ws.cell(row=row_idx, column=7).value
        
        if row_company and row_role and row_status in ["Lead", "Consideration"]:
            norm_co = normalize_string(row_company)
            if target_company and norm_co != normalize_string(target_company):
                continue
                
            # Check 1: Does it match search criteria and seniority levels?
            row_location = ws.cell(row=row_idx, column=5).value or "N/A"
            locs_list = [l.strip() for l in row_location.split(",") if l.strip()]
            
            if not matches_criteria(row_role, locs_list, search_criteria, row_company, config):
                print(f"Auto-archiving excluded criteria job: {row_company} - {row_role}")
                ws.cell(row=row_idx, column=7).value = "Archived"
                ws.cell(row=row_idx, column=1).value = "[ ]"
                ws.cell(row=row_idx, column=10).value = "Automatically archived: no longer matches search criteria / seniority levels."
                ws.cell(row=row_idx, column=15).value = "Not meeting Search Criteria"
                archived_criteria_count += 1
                continue
                
            # Check 2: Check if it's closed (only if the company was successfully scanned in this run)
            if norm_co in scanned_companies:
                key = (norm_co, normalize_string(row_role))
                if key not in crawled_keys:
                    # Double check if the URL is actually closed before archiving
                    cell = ws.cell(row=row_idx, column=8)
                    url = cell.hyperlink.target if cell.hyperlink else cell.value
                    
                    if url and not is_placeholder_url(url):
                        print(f"Double-checking if job is actually closed: {row_company} - {row_role} ({url})")
                        if is_job_url_closed(url, row_company):
                            print(f"Auto-archiving closed job: {row_company} - {row_role}")
                            ws.cell(row=row_idx, column=7).value = "Archived"
                            ws.cell(row=row_idx, column=1).value = "[ ]"
                            ws.cell(row=row_idx, column=10).value = "Automatically archived: job closed on careers portal."
                            ws.cell(row=row_idx, column=15).value = "Closed"
                            archived_closed_count += 1
                        else:
                            print(f"Keeping job active: {row_company} - {row_role} (URL is still open/active)")
                    else:
                        print(f"Auto-archiving closed job (placeholder URL): {row_company} - {row_role}")
                        ws.cell(row=row_idx, column=7).value = "Archived"
                        ws.cell(row=row_idx, column=1).value = "[ ]"
                        ws.cell(row=row_idx, column=10).value = "Automatically archived: job closed on careers portal."
                        ws.cell(row=row_idx, column=15).value = "Closed"
                        archived_closed_count += 1
                        
    wb.save(TRACKER_PATH)
    print(f"Scan finished. {added_count} roles added, {updated_count} URLs updated, {archived_closed_count} closed jobs archived, {archived_criteria_count} criteria-excluded jobs archived.")

if __name__ == '__main__':
    import sys
    target_company = None
    if len(sys.argv) > 1:
        target_company = sys.argv[1]
    run_daily_scan(target_company)
